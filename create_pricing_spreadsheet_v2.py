#!/usr/bin/env python3
"""
Flowker Pricing Analysis Spreadsheet Generator - V2 CORRIGIDO
Todas as referências revisadas e validadas
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def create_pricing_spreadsheet():
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Number formats
    currency_brl = 'R$ #,##0.00'
    currency_usd = '$ #,##0.00'
    percent_fmt = '0.0%'
    number_fmt = '#,##0'
    decimal_fmt = '$ #,##0.000000'
    decimal_brl_fmt = 'R$ #,##0.000000'

    # =========================================================================
    # ABA 1: PREMISSAS - Estrutura clara e organizada
    # =========================================================================
    ws = wb.active
    ws.title = "PREMISSAS"

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 45

    # MAPA DE CÉLULAS (para referência):
    # B4 = Taxa de câmbio
    # B8:B13 = Preços cloud USD
    # B14 = Total infra USD
    # B18:B20 = Operações por workflow
    # B24:B26 = Custos por operação USD
    # B30 = Volume workflows/mês (CLIENTE TÍPICO)
    # B34:B37 = Pricing tiers (base)
    # C34:C37 = Pricing tiers (por wf)
    # D34:D37 = Pricing tiers (limite)
    # B41 = Custo variável/wf USD (calculado)
    # B42 = Custo variável/wf BRL (calculado)

    row = 1
    ws['A1'] = "ANÁLISE DE PRICING - FLOWKER"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')

    ws['A2'] = "Células AMARELAS = Inputs editáveis | Células VERDES = Calculadas"
    ws['A2'].font = Font(italic=True, size=10)
    row = 4

    # === CÂMBIO ===
    ws[f'A{row}'] = "TAXA DE CÂMBIO"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    ws[f'A{row}'] = "USD para BRL"
    ws[f'B{row}'] = 5.00  # B5
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].border = thin_border
    ws[f'C{row}'] = "BRL/USD"
    ws[f'D{row}'] = "← Ajuste conforme câmbio atual"
    CAMBIO = 'B5'
    row += 2

    # === PREÇOS CLOUD ===
    ws[f'A{row}'] = "CUSTOS DE INFRAESTRUTURA (USD/mês) - STANDALONE"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1  # row = 8

    ws[f'A{row}'] = "Componente"
    ws[f'B{row}'] = "USD/mês"
    ws[f'C{row}'] = "BRL/mês"
    ws[f'D{row}'] = "Configuração"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}{row}'].fill = subheader_fill
    row += 1  # row = 9

    cloud = [
        ("PostgreSQL (RDS db.r6g.large)", 108, "2 vCPU, 16GB RAM"),        # B9
        ("MongoDB (Atlas M30)", 389, "2 vCPU, 8GB RAM"),                    # B10
        ("Valkey (ElastiCache cache.r6g.large)", 94, "2 vCPU, 13GB RAM"),  # B11
        ("Temporal Cloud (base)", 25, "Orquestração de workflows"),        # B12
        ("RabbitMQ (Amazon MQ mq.m5.large)", 180, "2 vCPU, 8GB RAM"),      # B13
        ("Compute (EKS 2x c6g.large)", 122, "4 vCPU total, 8GB RAM"),      # B14
    ]

    CLOUD_START = row  # 9
    for item in cloud:
        ws[f'A{row}'] = item[0]
        ws[f'B{row}'] = item[1]
        ws[f'B{row}'].fill = input_fill
        ws[f'B{row}'].border = thin_border
        ws[f'B{row}'].number_format = currency_usd
        ws[f'C{row}'] = f'=B{row}*{CAMBIO}'
        ws[f'C{row}'].fill = calc_fill
        ws[f'C{row}'].border = thin_border
        ws[f'C{row}'].number_format = currency_brl
        ws[f'D{row}'] = item[2]
        row += 1
    CLOUD_END = row - 1  # 14

    # Total infra
    ws[f'A{row}'] = "TOTAL CUSTO FIXO MENSAL"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f'=SUM(B{CLOUD_START}:B{CLOUD_END})'  # B15
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].fill = calc_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = currency_usd
    ws[f'C{row}'] = f'=B{row}*{CAMBIO}'  # C15
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'C{row}'].fill = calc_fill
    ws[f'C{row}'].border = thin_border
    ws[f'C{row}'].number_format = currency_brl
    TOTAL_FIXO_USD = f'B{row}'  # B15
    TOTAL_FIXO_BRL = f'C{row}'  # C15
    row += 2

    # === OPERAÇÕES POR WORKFLOW ===
    ws[f'A{row}'] = "OPERAÇÕES POR WORKFLOW (baseado em código)"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1  # 18

    ws[f'A{row}'] = "Operação"
    ws[f'B{row}'] = "Qtd/WF"
    ws[f'C{row}'] = "Unidade"
    ws[f'D{row}'] = "Fonte (arquivo)"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}{row}'].fill = subheader_fill
    row += 1  # 19

    ops = [
        ("Temporal Actions", 9, "actions", "validation_workflow.go"),     # B19
        ("MongoDB Operations", 6, "ops", "activities.go (audit)"),        # B20
        ("Valkey Operations", 12, "ops", "activities.go (locks/cache)"),  # B21
    ]

    OPS_ROWS = {}
    for item in ops:
        ws[f'A{row}'] = item[0]
        ws[f'B{row}'] = item[1]
        ws[f'B{row}'].fill = input_fill
        ws[f'B{row}'].border = thin_border
        ws[f'C{row}'] = item[2]
        ws[f'D{row}'] = item[3]
        OPS_ROWS[item[0]] = row
        row += 1

    TEMPORAL_OPS = f'B{OPS_ROWS["Temporal Actions"]}'      # B19
    MONGO_OPS = f'B{OPS_ROWS["MongoDB Operations"]}'       # B20
    VALKEY_OPS = f'B{OPS_ROWS["Valkey Operations"]}'       # B21
    row += 1

    # === CUSTOS POR OPERAÇÃO ===
    ws[f'A{row}'] = "CUSTOS POR OPERAÇÃO (USD)"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1  # 24

    ws[f'A{row}'] = "Operação"
    ws[f'B{row}'] = "USD/op"
    ws[f'C{row}'] = "BRL/op"
    ws[f'D{row}'] = "Fonte"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}{row}'].fill = subheader_fill
    row += 1  # 25

    costs = [
        ("Temporal (por action)", 0.000025, "Temporal Cloud Pricing"),  # B25
        ("MongoDB (por operação)", 0.000001, "Estimativa Atlas"),       # B26
        ("Valkey (por operação)", 0.0000001, "Estimativa ElastiCache"), # B27
    ]

    COST_ROWS = {}
    for item in costs:
        ws[f'A{row}'] = item[0]
        ws[f'B{row}'] = item[1]
        ws[f'B{row}'].fill = input_fill
        ws[f'B{row}'].border = thin_border
        ws[f'B{row}'].number_format = decimal_fmt
        ws[f'C{row}'] = f'=B{row}*{CAMBIO}'
        ws[f'C{row}'].fill = calc_fill
        ws[f'C{row}'].border = thin_border
        ws[f'C{row}'].number_format = decimal_brl_fmt
        ws[f'D{row}'] = item[2]
        COST_ROWS[item[0]] = row
        row += 1

    TEMPORAL_COST = f'B{COST_ROWS["Temporal (por action)"]}'      # B25
    MONGO_COST = f'B{COST_ROWS["MongoDB (por operação)"]}'        # B26
    VALKEY_COST = f'B{COST_ROWS["Valkey (por operação)"]}'        # B27
    row += 1

    # === VOLUME DE WORKFLOWS ===
    ws[f'A{row}'] = "VOLUME DE WORKFLOWS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1  # 30

    ws[f'A{row}'] = "Workflows/mês (cliente típico Growth)"
    ws[f'B{row}'] = 25000  # B30
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = number_fmt
    ws[f'D{row}'] = "← Volume médio esperado por cliente"
    VOLUME_WF = f'B{row}'  # B30
    row += 2

    # === PRICING TIERS ===
    ws[f'A{row}'] = "PRICING TIERS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1  # 33

    ws[f'A{row}'] = "Tier"
    ws[f'B{row}'] = "Base (BRL)"
    ws[f'C{row}'] = "Por WF (BRL)"
    ws[f'D{row}'] = "Limite WF/mês"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}{row}'].fill = subheader_fill
    row += 1  # 34

    tiers = [
        ("Starter", 0, 0, 1000),           # row 34
        ("Growth", 2990, 0.08, 25000),      # row 35
        ("Scale", 9990, 0.04, 200000),      # row 36
        ("Enterprise", 19990, 0.02, 1000000), # row 37
    ]

    TIER_ROWS = {}
    for tier in tiers:
        ws[f'A{row}'] = tier[0]
        ws[f'B{row}'] = tier[1]
        ws[f'B{row}'].fill = input_fill
        ws[f'B{row}'].border = thin_border
        ws[f'B{row}'].number_format = currency_brl
        ws[f'C{row}'] = tier[2]
        ws[f'C{row}'].fill = input_fill
        ws[f'C{row}'].border = thin_border
        ws[f'C{row}'].number_format = 'R$ #,##0.00'
        ws[f'D{row}'] = tier[3]
        ws[f'D{row}'].fill = input_fill
        ws[f'D{row}'].border = thin_border
        ws[f'D{row}'].number_format = number_fmt
        TIER_ROWS[tier[0]] = {'row': row, 'base': f'B{row}', 'per_wf': f'C{row}', 'limit': f'D{row}'}
        row += 1
    row += 1

    # === CÁLCULOS DERIVADOS ===
    ws[f'A{row}'] = "CÁLCULOS DERIVADOS (não editar)"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1  # 40

    # Custo variável por workflow
    ws[f'A{row}'] = "Custo Variável/WF (USD)"
    # = (9 * 0.000025) + (6 * 0.000001) + (12 * 0.0000001)
    ws[f'B{row}'] = f'=({TEMPORAL_OPS}*{TEMPORAL_COST})+({MONGO_OPS}*{MONGO_COST})+({VALKEY_OPS}*{VALKEY_COST})'
    ws[f'B{row}'].fill = calc_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = decimal_fmt
    CUSTO_VAR_USD = f'B{row}'  # B40
    row += 1

    ws[f'A{row}'] = "Custo Variável/WF (BRL)"
    ws[f'B{row}'] = f'={CUSTO_VAR_USD}*{CAMBIO}'
    ws[f'B{row}'].fill = calc_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = decimal_brl_fmt
    CUSTO_VAR_BRL = f'B{row}'  # B41
    row += 2

    # Referências rápidas
    ws[f'A{row}'] = "REFERÊNCIAS RÁPIDAS"
    ws[f'A{row}'].font = Font(bold=True, color="666666")
    row += 1
    ws[f'A{row}'] = f"Custo Fixo Mensal (BRL): ={TOTAL_FIXO_BRL}"
    ws[f'A{row}'].font = Font(size=9, color="666666")
    row += 1
    ws[f'A{row}'] = f"Custo Variável/WF (BRL): ={CUSTO_VAR_BRL}"
    ws[f'A{row}'].font = Font(size=9, color="666666")
    row += 1
    ws[f'A{row}'] = f"Volume típico: ={VOLUME_WF}"
    ws[f'A{row}'].font = Font(size=9, color="666666")

    # =========================================================================
    # ABA 2: CUSTOS - Detalhamento
    # =========================================================================
    ws2 = wb.create_sheet("CUSTOS")

    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 15

    ws2['A1'] = "DETALHAMENTO DE CUSTOS"
    ws2['A1'].font = Font(bold=True, size=16)
    row = 3

    # Custos Fixos
    ws2[f'A{row}'] = "CUSTOS FIXOS MENSAIS (Infraestrutura Standalone)"
    ws2[f'A{row}'].font = header_font
    ws2[f'A{row}'].fill = header_fill
    ws2.merge_cells(f'A{row}:D{row}')
    row += 1

    ws2[f'A{row}'] = "Componente"
    ws2[f'B{row}'] = "USD/mês"
    ws2[f'C{row}'] = "BRL/mês"
    ws2[f'D{row}'] = "% Total"
    for col in ['A', 'B', 'C', 'D']:
        ws2[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws2[f'{col}{row}'].fill = subheader_fill
    row += 1

    # Referencia os valores da aba PREMISSAS
    components = [
        ("PostgreSQL", "PREMISSAS!B9"),
        ("MongoDB", "PREMISSAS!B10"),
        ("Valkey", "PREMISSAS!B11"),
        ("Temporal", "PREMISSAS!B12"),
        ("RabbitMQ", "PREMISSAS!B13"),
        ("Compute", "PREMISSAS!B14"),
    ]

    START_CUSTOS = row
    for comp in components:
        ws2[f'A{row}'] = comp[0]
        ws2[f'B{row}'] = f'={comp[1]}'
        ws2[f'B{row}'].fill = calc_fill
        ws2[f'B{row}'].number_format = currency_usd
        ws2[f'C{row}'] = f'=B{row}*PREMISSAS!B5'
        ws2[f'C{row}'].fill = calc_fill
        ws2[f'C{row}'].number_format = currency_brl
        ws2[f'D{row}'] = f'=B{row}/PREMISSAS!B15'
        ws2[f'D{row}'].fill = calc_fill
        ws2[f'D{row}'].number_format = percent_fmt
        row += 1
    END_CUSTOS = row - 1

    ws2[f'A{row}'] = "TOTAL"
    ws2[f'A{row}'].font = Font(bold=True)
    ws2[f'B{row}'] = f'=SUM(B{START_CUSTOS}:B{END_CUSTOS})'
    ws2[f'B{row}'].font = Font(bold=True)
    ws2[f'B{row}'].fill = calc_fill
    ws2[f'B{row}'].number_format = currency_usd
    ws2[f'C{row}'] = f'=SUM(C{START_CUSTOS}:C{END_CUSTOS})'
    ws2[f'C{row}'].font = Font(bold=True)
    ws2[f'C{row}'].fill = calc_fill
    ws2[f'C{row}'].number_format = currency_brl
    ws2[f'D{row}'] = '=100%'
    ws2[f'D{row}'].fill = calc_fill
    TOTAL_CUSTOS_ROW = row
    row += 2

    # Custos Variáveis
    ws2[f'A{row}'] = "CUSTOS VARIÁVEIS (por Workflow)"
    ws2[f'A{row}'].font = header_font
    ws2[f'A{row}'].fill = header_fill
    ws2.merge_cells(f'A{row}:D{row}')
    row += 1

    ws2[f'A{row}'] = "Componente"
    ws2[f'B{row}'] = "Ops/WF"
    ws2[f'C{row}'] = "Custo/Op (USD)"
    ws2[f'D{row}'] = "Custo/WF (USD)"
    for col in ['A', 'B', 'C', 'D']:
        ws2[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws2[f'{col}{row}'].fill = subheader_fill
    row += 1

    var_items = [
        ("Temporal Actions", "PREMISSAS!B19", "PREMISSAS!B25"),
        ("MongoDB Operations", "PREMISSAS!B20", "PREMISSAS!B26"),
        ("Valkey Operations", "PREMISSAS!B21", "PREMISSAS!B27"),
    ]

    START_VAR = row
    for item in var_items:
        ws2[f'A{row}'] = item[0]
        ws2[f'B{row}'] = f'={item[1]}'
        ws2[f'B{row}'].fill = calc_fill
        ws2[f'C{row}'] = f'={item[2]}'
        ws2[f'C{row}'].fill = calc_fill
        ws2[f'C{row}'].number_format = decimal_fmt
        ws2[f'D{row}'] = f'=B{row}*C{row}'
        ws2[f'D{row}'].fill = calc_fill
        ws2[f'D{row}'].number_format = decimal_fmt
        row += 1
    END_VAR = row - 1

    ws2[f'A{row}'] = "TOTAL CUSTO/WF (USD)"
    ws2[f'A{row}'].font = Font(bold=True)
    ws2[f'D{row}'] = f'=SUM(D{START_VAR}:D{END_VAR})'
    ws2[f'D{row}'].font = Font(bold=True)
    ws2[f'D{row}'].fill = calc_fill
    ws2[f'D{row}'].number_format = decimal_fmt
    TOTAL_VAR_USD_ROW = row
    row += 1

    ws2[f'A{row}'] = "TOTAL CUSTO/WF (BRL)"
    ws2[f'A{row}'].font = Font(bold=True)
    ws2[f'D{row}'] = f'=D{TOTAL_VAR_USD_ROW}*PREMISSAS!B5'
    ws2[f'D{row}'].font = Font(bold=True)
    ws2[f'D{row}'].fill = calc_fill
    ws2[f'D{row}'].number_format = decimal_brl_fmt
    TOTAL_VAR_BRL_ROW = row

    # =========================================================================
    # ABA 3: BREAK-EVEN
    # =========================================================================
    ws3 = wb.create_sheet("BREAK_EVEN")

    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 18
    ws3.column_dimensions['E'].width = 15

    ws3['A1'] = "ANÁLISE DE BREAK-EVEN (Standalone)"
    ws3['A1'].font = Font(bold=True, size=16)

    ws3['A2'] = "Fórmula: Break-even = Custo Fixo / (Preço/WF - Custo Variável/WF)"
    ws3['A2'].font = Font(italic=True)
    row = 4

    # Referências
    ws3[f'A{row}'] = "Custo Fixo Mensal (BRL):"
    ws3[f'B{row}'] = '=PREMISSAS!C15'
    ws3[f'B{row}'].fill = calc_fill
    ws3[f'B{row}'].number_format = currency_brl
    row += 1

    ws3[f'A{row}'] = "Custo Variável/WF (BRL):"
    ws3[f'B{row}'] = '=PREMISSAS!B41'
    ws3[f'B{row}'].fill = calc_fill
    ws3[f'B{row}'].number_format = decimal_brl_fmt
    CUSTO_VAR_REF = f'B{row}'
    row += 1

    ws3[f'A{row}'] = "Volume típico/cliente:"
    ws3[f'B{row}'] = '=PREMISSAS!B30'
    ws3[f'B{row}'].fill = calc_fill
    ws3[f'B{row}'].number_format = number_fmt
    VOLUME_REF = f'B{row}'
    row += 2

    # Tabela de break-even
    ws3[f'A{row}'] = "BREAK-EVEN POR PREÇO"
    ws3[f'A{row}'].font = header_font
    ws3[f'A{row}'].fill = header_fill
    ws3.merge_cells(f'A{row}:E{row}')
    row += 1

    ws3[f'A{row}'] = "Preço/WF"
    ws3[f'B{row}'] = "Margem/WF"
    ws3[f'C{row}'] = "Break-even (WF)"
    ws3[f'D{row}'] = "Clientes*"
    ws3[f'E{row}'] = "Viabilidade"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws3[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws3[f'{col}{row}'].fill = subheader_fill
    row += 1

    prices = [0.05, 0.08, 0.10, 0.15, 0.20]
    for price in prices:
        ws3[f'A{row}'] = price
        ws3[f'A{row}'].fill = input_fill
        ws3[f'A{row}'].border = thin_border
        ws3[f'A{row}'].number_format = 'R$ #,##0.00'
        # Margem = Preço - Custo Var
        ws3[f'B{row}'] = f'=A{row}-{CUSTO_VAR_REF}'
        ws3[f'B{row}'].fill = calc_fill
        ws3[f'B{row}'].border = thin_border
        ws3[f'B{row}'].number_format = 'R$ #,##0.0000'
        # Break-even = Custo Fixo / Margem
        ws3[f'C{row}'] = f'=IF(B{row}>0,B4/B{row},"N/A")'
        ws3[f'C{row}'].fill = calc_fill
        ws3[f'C{row}'].border = thin_border
        ws3[f'C{row}'].number_format = number_fmt
        # Clientes = Break-even / Volume típico
        ws3[f'D{row}'] = f'=IF(ISNUMBER(C{row}),C{row}/{VOLUME_REF},"N/A")'
        ws3[f'D{row}'].fill = calc_fill
        ws3[f'D{row}'].border = thin_border
        ws3[f'D{row}'].number_format = '#,##0.0'
        # Viabilidade
        ws3[f'E{row}'] = f'=IF(D{row}<=1,"Excelente",IF(D{row}<=2,"Bom",IF(D{row}<=4,"Aceitavel","Arriscado")))'
        ws3[f'E{row}'].fill = calc_fill
        ws3[f'E{row}'].border = thin_border
        row += 1

    row += 1
    ws3[f'A{row}'] = "*Clientes = Break-even / Volume típico por cliente"
    ws3[f'A{row}'].font = Font(italic=True, size=9)

    # =========================================================================
    # ABA 4: SIMULADOR (Standalone only)
    # =========================================================================
    ws4 = wb.create_sheet("SIMULADOR")

    ws4.column_dimensions['A'].width = 35
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 20
    ws4.column_dimensions['D'].width = 25

    ws4['A1'] = "SIMULADOR DE MARGEM (Standalone)"
    ws4['A1'].font = Font(bold=True, size=16)

    ws4['A2'] = "Todos os cálculos assumem infraestrutura 100% dedicada"
    ws4['A2'].font = Font(italic=True)
    row = 4

    # Inputs do simulador
    ws4[f'A{row}'] = "INPUTS DO SIMULADOR"
    ws4[f'A{row}'].font = header_font
    ws4[f'A{row}'].fill = header_fill
    ws4.merge_cells(f'A{row}:D{row}')
    row += 1

    ws4[f'A{row}'] = "Preço Base (BRL):"
    ws4[f'B{row}'] = 2990
    ws4[f'B{row}'].fill = input_fill
    ws4[f'B{row}'].border = thin_border
    ws4[f'B{row}'].number_format = currency_brl
    ws4[f'D{row}'] = "← Ajuste o preço base"
    SIM_BASE = f'B{row}'
    row += 1

    ws4[f'A{row}'] = "Preço por Workflow (BRL):"
    ws4[f'B{row}'] = 0.08
    ws4[f'B{row}'].fill = input_fill
    ws4[f'B{row}'].border = thin_border
    ws4[f'B{row}'].number_format = 'R$ #,##0.00'
    ws4[f'D{row}'] = "← Ajuste o preço por workflow"
    SIM_PER_WF = f'B{row}'
    row += 1

    ws4[f'A{row}'] = "Volume de Workflows/mês:"
    ws4[f'B{row}'] = 25000
    ws4[f'B{row}'].fill = input_fill
    ws4[f'B{row}'].border = thin_border
    ws4[f'B{row}'].number_format = number_fmt
    ws4[f'D{row}'] = "← Ajuste o volume esperado"
    SIM_VOL = f'B{row}'
    row += 2

    # Resultados
    ws4[f'A{row}'] = "RESULTADOS"
    ws4[f'A{row}'].font = header_font
    ws4[f'A{row}'].fill = header_fill
    ws4.merge_cells(f'A{row}:D{row}')
    row += 1

    ws4[f'A{row}'] = "Receita Base:"
    ws4[f'B{row}'] = f'={SIM_BASE}'
    ws4[f'B{row}'].fill = calc_fill
    ws4[f'B{row}'].number_format = currency_brl
    row += 1

    ws4[f'A{row}'] = "Receita Variável:"
    ws4[f'B{row}'] = f'={SIM_VOL}*{SIM_PER_WF}'
    ws4[f'B{row}'].fill = calc_fill
    ws4[f'B{row}'].number_format = currency_brl
    SIM_REC_VAR = f'B{row}'
    row += 1

    ws4[f'A{row}'] = "RECEITA TOTAL:"
    ws4[f'A{row}'].font = Font(bold=True)
    ws4[f'B{row}'] = f'={SIM_BASE}+{SIM_REC_VAR}'
    ws4[f'B{row}'].font = Font(bold=True)
    ws4[f'B{row}'].fill = calc_fill
    ws4[f'B{row}'].number_format = currency_brl
    SIM_RECEITA = f'B{row}'
    row += 2

    ws4[f'A{row}'] = "Custo Fixo (infra standalone):"
    ws4[f'B{row}'] = '=PREMISSAS!C15'
    ws4[f'B{row}'].fill = calc_fill
    ws4[f'B{row}'].number_format = currency_brl
    SIM_FIXO = f'B{row}'
    row += 1

    ws4[f'A{row}'] = "Custo Variável:"
    ws4[f'B{row}'] = f'={SIM_VOL}*PREMISSAS!B41'
    ws4[f'B{row}'].fill = calc_fill
    ws4[f'B{row}'].number_format = currency_brl
    SIM_VAR = f'B{row}'
    row += 1

    ws4[f'A{row}'] = "CUSTO TOTAL:"
    ws4[f'A{row}'].font = Font(bold=True)
    ws4[f'B{row}'] = f'={SIM_FIXO}+{SIM_VAR}'
    ws4[f'B{row}'].font = Font(bold=True)
    ws4[f'B{row}'].fill = calc_fill
    ws4[f'B{row}'].number_format = currency_brl
    SIM_CUSTO = f'B{row}'
    row += 2

    ws4[f'A{row}'] = "MARGEM BRUTA (R$):"
    ws4[f'A{row}'].font = Font(bold=True, size=14)
    ws4[f'B{row}'] = f'={SIM_RECEITA}-{SIM_CUSTO}'
    ws4[f'B{row}'].font = Font(bold=True, size=14)
    ws4[f'B{row}'].fill = calc_fill
    ws4[f'B{row}'].number_format = currency_brl
    SIM_MARGEM = f'B{row}'
    row += 1

    ws4[f'A{row}'] = "MARGEM BRUTA (%):"
    ws4[f'A{row}'].font = Font(bold=True, size=14)
    ws4[f'B{row}'] = f'=IF({SIM_RECEITA}>0,{SIM_MARGEM}/{SIM_RECEITA},0)'
    ws4[f'B{row}'].font = Font(bold=True, size=14)
    ws4[f'B{row}'].fill = calc_fill
    ws4[f'B{row}'].number_format = percent_fmt
    row += 2

    ws4[f'A{row}'] = "STATUS:"
    ws4[f'A{row}'].font = Font(bold=True, size=14)
    ws4[f'B{row}'] = f'=IF({SIM_MARGEM}>0,"SUSTENTAVEL","PREJUIZO")'
    ws4[f'B{row}'].font = Font(bold=True, size=14)
    ws4[f'B{row}'].fill = calc_fill

    # =========================================================================
    # ABA 5: CONFIANÇA
    # =========================================================================
    ws5 = wb.create_sheet("CONFIANCA")

    ws5.column_dimensions['A'].width = 40
    ws5.column_dimensions['B'].width = 18
    ws5.column_dimensions['C'].width = 50

    ws5['A1'] = "GRAUS DE CONFIANÇA"
    ws5['A1'].font = Font(bold=True, size=16)
    row = 3

    ws5[f'A{row}'] = "Componente"
    ws5[f'B{row}'] = "Confiança"
    ws5[f'C{row}'] = "Justificativa"
    for col in ['A', 'B', 'C']:
        ws5[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws5[f'{col}{row}'].fill = header_fill
    row += 1

    items = [
        ("Metodologia standalone-first", "ALTO", "Princípio de negócio conservador"),
        ("Componentes de infraestrutura", "ALTO", "Baseado em docker-compose.yml real"),
        ("Preços de cloud (unitários)", "ALTO", "Pricing público AWS/MongoDB/Temporal"),
        ("Sizing de produção", "BAIXO-MEDIO", "Sem Terraform - estimar conservador"),
        ("Operações por workflow", "BAIXO-MEDIO", "Baseado em código, não em métricas"),
        ("Custo variável por workflow", "BAIXO-MEDIO", "Requer validação com benchmark"),
        ("Modelo de pricing (estrutura)", "ALTO", "Padrão de mercado (base + variável)"),
        ("Modelo de pricing (valores)", "MEDIO", "Depende de validação de custos"),
    ]

    for item in items:
        ws5[f'A{row}'] = item[0]
        ws5[f'B{row}'] = item[1]
        ws5[f'C{row}'] = item[2]
        row += 1

    row += 2
    ws5[f'A{row}'] = "AÇÕES PARA AUMENTAR CONFIANÇA:"
    ws5[f'A{row}'].font = Font(bold=True)
    row += 1

    acoes = [
        "1. Executar benchmark script com Docker rodando",
        "2. Instalar KubeCost no cluster K8s",
        "3. Rodar load test com 25K-100K workflows",
        "4. Criar Terraform do Flowker para produção",
        "5. Coletar 30+ dias de dados de billing real",
    ]
    for acao in acoes:
        ws5[f'A{row}'] = acao
        row += 1

    # Save
    output_path = "/tmp/pricing-flowker/Flowker_Pricing_Model_v2.xlsx"
    wb.save(output_path)
    print(f"Planilha criada: {output_path}")
    return output_path

if __name__ == "__main__":
    create_pricing_spreadsheet()
