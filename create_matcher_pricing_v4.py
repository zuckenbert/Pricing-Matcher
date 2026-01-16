#!/usr/bin/env python3
"""
Matcher Pricing Model v4 - SIMPLIFICADO (igual Flowker)
========================================================

Abordagem simples:
1. Defino UMA configuração de infra
2. Calculo o custo dessa infra
3. Estimo a capacidade máxima (QPS → TPS → txns/mês)
4. Defino preço com margem

Sem múltiplos tiers complicados - igual fiz no Flowker.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CRIAÇÃO DA PLANILHA
# ============================================================================

wb = Workbook()

# Estilos
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
section_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
section_font = Font(bold=True, size=11)
formula_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
warning_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
result_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(cell):
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

def style_section(cell):
    cell.fill = section_fill
    cell.font = section_font
    cell.border = thin_border

def style_cell(cell):
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='left', vertical='center')

def style_number(cell):
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = '#,##0'

def style_decimal(cell):
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = '#,##0.00'

def style_percent(cell):
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = '0.0%'

def style_currency_brl(cell):
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = '"R$" #,##0'

def style_currency_usd(cell):
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = '"$" #,##0'

# ============================================================================
# ABA 1: MODELO (tudo junto)
# ============================================================================

ws = wb.active
ws.title = "MODELO"

row = 1

# Título
ws['A1'] = "PRICING MATCHER - MODELO SIMPLIFICADO"
ws['A1'].font = Font(bold=True, size=16)
ws.merge_cells('A1:E1')
row = 3

# ============================================================================
# SEÇÃO 1: PARÂMETROS GERAIS
# ============================================================================

ws[f'A{row}'] = "1. PARÂMETROS GERAIS"
style_section(ws[f'A{row}'])
ws.merge_cells(f'A{row}:E{row}')
row += 1

headers = ["Parâmetro", "Valor", "Unidade", "Descrição", "Confiança"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    style_header(cell)
row += 1

params = [
    ("Câmbio", 5.0, "BRL/USD", "Taxa de conversão", "100%"),
    ("Ops Redis por transação", 5, "ops", "dedupe + locks (análise de código)", "75%"),
    ("Ops PostgreSQL por transação", 6, "ops", "exists + insert + select + match", "70%"),
    ("Fator de pico", 3, "x", "TPS pico = TPS médio × fator", "50%"),
    ("Segundos no mês", 2592000, "seg", "30 dias × 24h × 3600s", "100%"),
    ("Margem alvo", 0.70, "%", "Margem bruta mínima", "N/A"),
]

param_rows = {}
for p in params:
    ws.cell(row=row, column=1, value=p[0])
    style_cell(ws.cell(row=row, column=1))
    ws.cell(row=row, column=2, value=p[1])
    if p[0] == "Margem alvo":
        style_percent(ws.cell(row=row, column=2))
    elif isinstance(p[1], float):
        style_decimal(ws.cell(row=row, column=2))
    else:
        style_number(ws.cell(row=row, column=2))
    ws.cell(row=row, column=2).fill = input_fill
    ws.cell(row=row, column=3, value=p[2])
    style_cell(ws.cell(row=row, column=3))
    ws.cell(row=row, column=4, value=p[3])
    style_cell(ws.cell(row=row, column=4))
    ws.cell(row=row, column=5, value=p[4])
    style_cell(ws.cell(row=row, column=5))
    param_rows[p[0]] = row
    row += 1

# Total ops/transação (calculado)
ws.cell(row=row, column=1, value="TOTAL Ops por transação")
style_cell(ws.cell(row=row, column=1))
ws.cell(row=row, column=1).font = Font(bold=True)
ws.cell(row=row, column=2, value=f"=B{param_rows['Ops Redis por transação']}+B{param_rows['Ops PostgreSQL por transação']}")
style_number(ws.cell(row=row, column=2))
ws.cell(row=row, column=2).fill = formula_fill
ws.cell(row=row, column=3, value="ops/txn")
style_cell(ws.cell(row=row, column=3))
ws.cell(row=row, column=4, value="= 1 TPS gera X QPS")
style_cell(ws.cell(row=row, column=4))
ws.cell(row=row, column=5, value="70%")
style_cell(ws.cell(row=row, column=5))
total_ops_row = row
row += 2

# ============================================================================
# SEÇÃO 2: INFRAESTRUTURA E CUSTOS
# ============================================================================

ws[f'A{row}'] = "2. INFRAESTRUTURA E CUSTOS (configuração única)"
style_section(ws[f'A{row}'])
ws.merge_cells(f'A{row}:E{row}')
row += 1

headers = ["Componente", "Sizing", "Custo USD", "Custo BRL", "Confiança"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    style_header(cell)
row += 1

componentes = [
    ("PostgreSQL", "db.t3.large (2 vCPU, 8GB)", 160, "100%"),
    ("Redis", "cache.t3.medium (2 vCPU, 3GB)", 50, "100%"),
    ("RabbitMQ", "mq.t3.micro (2 vCPU, 1GB)", 59, "100%"),
    ("Fargate (API + Workers)", "5x (1 vCPU, 2GB cada)", 162, "100%"),
    ("Outros (rede, logs, S3)", "Estimado", 70, "60%"),
]

comp_start_row = row
for comp in componentes:
    ws.cell(row=row, column=1, value=comp[0])
    style_cell(ws.cell(row=row, column=1))
    ws.cell(row=row, column=2, value=comp[1])
    style_cell(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3, value=comp[2])
    style_currency_usd(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).fill = input_fill
    # Custo BRL = USD × Câmbio
    ws.cell(row=row, column=4, value=f"=C{row}*$B${param_rows['Câmbio']}")
    style_currency_brl(ws.cell(row=row, column=4))
    ws.cell(row=row, column=4).fill = formula_fill
    ws.cell(row=row, column=5, value=comp[3])
    style_cell(ws.cell(row=row, column=5))
    row += 1
comp_end_row = row - 1

# Total
ws.cell(row=row, column=1, value="TOTAL CUSTO INFRA")
style_cell(ws.cell(row=row, column=1))
ws.cell(row=row, column=1).font = Font(bold=True)
ws.cell(row=row, column=2, value="")
style_cell(ws.cell(row=row, column=2))
ws.cell(row=row, column=3, value=f"=SUM(C{comp_start_row}:C{comp_end_row})")
style_currency_usd(ws.cell(row=row, column=3))
ws.cell(row=row, column=3).fill = formula_fill
ws.cell(row=row, column=3).font = Font(bold=True)
ws.cell(row=row, column=4, value=f"=SUM(D{comp_start_row}:D{comp_end_row})")
style_currency_brl(ws.cell(row=row, column=4))
ws.cell(row=row, column=4).fill = formula_fill
ws.cell(row=row, column=4).font = Font(bold=True)
ws.cell(row=row, column=5, value="95%")
style_cell(ws.cell(row=row, column=5))
custo_total_row = row
row += 2

# ============================================================================
# SEÇÃO 3: CAPACIDADE ESTIMADA
# ============================================================================

ws[f'A{row}'] = "3. CAPACIDADE ESTIMADA (sem teste de carga!)"
style_section(ws[f'A{row}'])
ws.merge_cells(f'A{row}:E{row}')
row += 1

headers = ["Métrica", "Valor", "Unidade", "Fórmula", "Confiança"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    style_header(cell)
row += 1

# QPS máximo (input - estimativa baseada em specs)
ws.cell(row=row, column=1, value="QPS máximo estimado")
style_cell(ws.cell(row=row, column=1))
ws.cell(row=row, column=2, value=100)  # Estimativa para db.t3.large
style_number(ws.cell(row=row, column=2))
ws.cell(row=row, column=2).fill = input_fill
ws.cell(row=row, column=3, value="QPS")
style_cell(ws.cell(row=row, column=3))
ws.cell(row=row, column=4, value="Estimativa para db.t3.large")
style_cell(ws.cell(row=row, column=4))
ws.cell(row=row, column=5, value="55%")
style_cell(ws.cell(row=row, column=5))
ws.cell(row=row, column=5).fill = warning_fill
qps_row = row
row += 1

# TPS máximo (calculado)
ws.cell(row=row, column=1, value="TPS máximo")
style_cell(ws.cell(row=row, column=1))
ws.cell(row=row, column=2, value=f"=B{qps_row}/B{total_ops_row}")
style_decimal(ws.cell(row=row, column=2))
ws.cell(row=row, column=2).fill = formula_fill
ws.cell(row=row, column=3, value="TPS")
style_cell(ws.cell(row=row, column=3))
ws.cell(row=row, column=4, value="= QPS ÷ Ops/txn")
style_cell(ws.cell(row=row, column=4))
ws.cell(row=row, column=5, value="55%")
style_cell(ws.cell(row=row, column=5))
tps_row = row
row += 1

# Transações/mês máximo (calculado considerando pico)
ws.cell(row=row, column=1, value="Transações/mês máximo")
style_cell(ws.cell(row=row, column=1))
ws.cell(row=row, column=2, value=f"=(B{tps_row}/B{param_rows['Fator de pico']})*B{param_rows['Segundos no mês']}")
style_number(ws.cell(row=row, column=2))
ws.cell(row=row, column=2).fill = formula_fill
ws.cell(row=row, column=3, value="txns/mês")
style_cell(ws.cell(row=row, column=3))
ws.cell(row=row, column=4, value="= (TPS ÷ Fator_pico) × Seg_mês")
style_cell(ws.cell(row=row, column=4))
ws.cell(row=row, column=5, value="50%")
style_cell(ws.cell(row=row, column=5))
ws.cell(row=row, column=5).fill = warning_fill
txns_row = row
row += 2

# ============================================================================
# SEÇÃO 4: PRICING
# ============================================================================

ws[f'A{row}'] = "4. PRICING SUGERIDO"
style_section(ws[f'A{row}'])
ws.merge_cells(f'A{row}:E{row}')
row += 1

headers = ["Item", "Valor", "", "Fórmula", ""]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    style_header(cell)
row += 1

# Custo
ws.cell(row=row, column=1, value="Custo mensal")
style_cell(ws.cell(row=row, column=1))
ws.cell(row=row, column=2, value=f"=D{custo_total_row}")
style_currency_brl(ws.cell(row=row, column=2))
ws.cell(row=row, column=2).fill = formula_fill
ws.cell(row=row, column=4, value="= Total infra")
style_cell(ws.cell(row=row, column=4))
custo_pricing_row = row
row += 1

# Preço sugerido
ws.cell(row=row, column=1, value="Preço sugerido")
style_cell(ws.cell(row=row, column=1))
ws.cell(row=row, column=1).font = Font(bold=True, size=12)
ws.cell(row=row, column=2, value=f"=B{custo_pricing_row}/(1-B{param_rows['Margem alvo']})")
style_currency_brl(ws.cell(row=row, column=2))
ws.cell(row=row, column=2).fill = result_fill
ws.cell(row=row, column=2).font = Font(bold=True, size=14)
ws.cell(row=row, column=4, value="= Custo ÷ (1 - Margem)")
style_cell(ws.cell(row=row, column=4))
preco_row = row
row += 1

# Margem
ws.cell(row=row, column=1, value="Margem resultante")
style_cell(ws.cell(row=row, column=1))
ws.cell(row=row, column=2, value=f"=(B{preco_row}-B{custo_pricing_row})/B{preco_row}")
style_percent(ws.cell(row=row, column=2))
ws.cell(row=row, column=2).fill = formula_fill
ws.cell(row=row, column=4, value="= (Preço - Custo) / Preço")
style_cell(ws.cell(row=row, column=4))
row += 1

# Volume incluso
ws.cell(row=row, column=1, value="Volume incluso (até)")
style_cell(ws.cell(row=row, column=1))
ws.cell(row=row, column=2, value=f"=B{txns_row}")
style_number(ws.cell(row=row, column=2))
ws.cell(row=row, column=2).fill = formula_fill
ws.cell(row=row, column=3, value="txns/mês")
style_cell(ws.cell(row=row, column=3))
ws.cell(row=row, column=4, value="= Capacidade máxima")
style_cell(ws.cell(row=row, column=4))
row += 2

# ============================================================================
# SEÇÃO 5: RESUMO
# ============================================================================

ws[f'A{row}'] = "5. RESUMO"
style_section(ws[f'A{row}'])
ws.merge_cells(f'A{row}:E{row}')
row += 1

ws.cell(row=row, column=1, value="PREÇO TETO:")
ws.cell(row=row, column=1).font = Font(bold=True, size=12)
ws.cell(row=row, column=2, value=f"=B{preco_row}")
style_currency_brl(ws.cell(row=row, column=2))
ws.cell(row=row, column=2).fill = result_fill
ws.cell(row=row, column=2).font = Font(bold=True, size=14)
ws.cell(row=row, column=3, value="/mês")
row += 1

ws.cell(row=row, column=1, value="CAPACIDADE:")
ws.cell(row=row, column=1).font = Font(bold=True)
ws.cell(row=row, column=2, value=f"=B{txns_row}")
style_number(ws.cell(row=row, column=2))
ws.cell(row=row, column=2).fill = result_fill
ws.cell(row=row, column=3, value="txns/mês")
row += 1

ws.cell(row=row, column=1, value="CONFIANÇA:")
ws.cell(row=row, column=1).font = Font(bold=True)
ws.cell(row=row, column=2, value="65%")
ws.cell(row=row, column=2).fill = warning_fill
ws.cell(row=row, column=2).font = Font(bold=True)
ws.cell(row=row, column=3, value="±35%")
row += 2

# Aviso
ws[f'A{row}'] = "ATENÇÃO: Este é um PREÇO TETO baseado em infra dedicada (standalone)."
ws[f'A{row}'].font = Font(italic=True, color="CC0000")
ws.merge_cells(f'A{row}:E{row}')
row += 1
ws[f'A{row}'] = "Capacidade NÃO validada por teste de carga. Confiança: 65% ±35%."
ws[f'A{row}'].font = Font(italic=True, color="CC0000")
ws.merge_cells(f'A{row}:E{row}')

# Ajustar larguras
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 12

# ============================================================================
# ABA 2: CONFIANÇA (simplificada)
# ============================================================================

ws2 = wb.create_sheet("CONFIANCA")

ws2['A1'] = "NÍVEIS DE CONFIANÇA"
ws2['A1'].font = Font(bold=True, size=14)
ws2.merge_cells('A1:C1')

ws2['A3'] = "Componente"
ws2['B3'] = "Confiança"
ws2['C3'] = "O que validaria"
for col in ['A', 'B', 'C']:
    style_header(ws2[f'{col}3'])

confiancas = [
    ("Preços AWS", "100%", "Nada - são públicos"),
    ("Componentes de infra", "100%", "Nada - vem do docker-compose"),
    ("Ops/transação (11)", "70%", "Profiling em produção"),
    ("QPS capacidade (100)", "55%", "Teste de carga"),
    ("Fator de pico (3x)", "50%", "Dados de produção"),
    ("GERAL", "65% ±35%", "Teste de carga + KubeCost"),
]

for i, (comp, conf, valid) in enumerate(confiancas, start=4):
    ws2[f'A{i}'] = comp
    style_cell(ws2[f'A{i}'])
    if comp == "GERAL":
        ws2[f'A{i}'].font = Font(bold=True)
    ws2[f'B{i}'] = conf
    style_cell(ws2[f'B{i}'])
    if comp == "GERAL":
        ws2[f'B{i}'].fill = warning_fill
        ws2[f'B{i}'].font = Font(bold=True)
    ws2[f'C{i}'] = valid
    style_cell(ws2[f'C{i}'])

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 30

# ============================================================================
# SALVAR
# ============================================================================

output_path = "/private/tmp/pricing-flowker/Matcher_Pricing_Model_v4.xlsx"
wb.save(output_path)

print("=" * 60)
print("PLANILHA MATCHER V4 - SIMPLIFICADA")
print("=" * 60)
print(f"\nArquivo: {output_path}")
print("\nESTRUTURA:")
print("  1. MODELO - Tudo junto (params + infra + capacidade + pricing)")
print("  2. CONFIANCA - Níveis de confiança")
print("\nLÓGICA (igual Flowker):")
print("  1 config de infra → custo → capacidade → preço")
print("=" * 60)
