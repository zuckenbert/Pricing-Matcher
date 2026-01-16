#!/usr/bin/env python3
"""
Flowker Pricing Analysis Spreadsheet Generator
Creates an Excel file with editable assumptions and linked formulas
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.comments import Comment

def create_pricing_spreadsheet():
    wb = Workbook()

    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Yellow for inputs
    calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # Green for calcs
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    currency_format = 'R$ #,##0.00'
    currency_usd_format = '$ #,##0.00'
    percent_format = '0.0%'
    number_format = '#,##0'
    decimal_format = '#,##0.000000'

    # =========================================================================
    # ABA 1: PREMISSAS (Inputs)
    # =========================================================================
    ws_premissas = wb.active
    ws_premissas.title = "PREMISSAS"

    # Column widths
    ws_premissas.column_dimensions['A'].width = 40
    ws_premissas.column_dimensions['B'].width = 20
    ws_premissas.column_dimensions['C'].width = 15
    ws_premissas.column_dimensions['D'].width = 40

    row = 1

    # Title
    ws_premissas['A1'] = "ANÁLISE DE PRICING - FLOWKER"
    ws_premissas['A1'].font = Font(bold=True, size=16)
    ws_premissas.merge_cells('A1:D1')

    ws_premissas['A2'] = "Premissas Editáveis (células amarelas)"
    ws_premissas['A2'].font = Font(italic=True, size=10)
    row = 4

    # ----- SEÇÃO: CÂMBIO -----
    ws_premissas[f'A{row}'] = "TAXA DE CÂMBIO"
    ws_premissas[f'A{row}'].font = header_font
    ws_premissas[f'A{row}'].fill = header_fill
    ws_premissas.merge_cells(f'A{row}:D{row}')
    row += 1

    ws_premissas[f'A{row}'] = "USD para BRL"
    ws_premissas[f'B{row}'] = 5.00
    ws_premissas[f'B{row}'].fill = input_fill
    ws_premissas[f'B{row}'].border = thin_border
    ws_premissas[f'C{row}'] = "BRL/USD"
    ws_premissas[f'D{row}'] = "← EDITÁVEL: Ajuste conforme câmbio atual"
    CAMBIO_CELL = f'B{row}'  # B5
    row += 2

    # ----- SEÇÃO: PREÇOS CLOUD (USD) -----
    ws_premissas[f'A{row}'] = "PREÇOS DE CLOUD (USD/mês)"
    ws_premissas[f'A{row}'].font = header_font
    ws_premissas[f'A{row}'].fill = header_fill
    ws_premissas.merge_cells(f'A{row}:D{row}')
    row += 1

    # Subheaders
    ws_premissas[f'A{row}'] = "Componente"
    ws_premissas[f'B{row}'] = "Preço USD"
    ws_premissas[f'C{row}'] = "Preço BRL"
    ws_premissas[f'D{row}'] = "Configuração"
    for col in ['A', 'B', 'C', 'D']:
        ws_premissas[f'{col}{row}'].font = Font(bold=True)
        ws_premissas[f'{col}{row}'].fill = subheader_fill
        ws_premissas[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
    row += 1

    cloud_prices = [
        ("PostgreSQL (RDS)", 108, "db.r6g.large (2 vCPU, 16GB)"),
        ("MongoDB (Atlas)", 389, "M30 (2 vCPU, 8GB)"),
        ("Valkey (ElastiCache)", 94, "cache.r6g.large (2 vCPU, 13GB)"),
        ("Temporal Cloud", 25, "Base plan"),
        ("RabbitMQ (Amazon MQ)", 180, "mq.m5.large (2 vCPU, 8GB)"),
        ("Compute (EKS)", 122, "2x c6g.large (2 vCPU, 4GB each)"),
    ]

    cloud_price_cells = {}
    CLOUD_START_ROW = row
    for item in cloud_prices:
        ws_premissas[f'A{row}'] = item[0]
        ws_premissas[f'B{row}'] = item[1]
        ws_premissas[f'B{row}'].fill = input_fill
        ws_premissas[f'B{row}'].border = thin_border
        ws_premissas[f'B{row}'].number_format = currency_usd_format
        # Formula for BRL conversion
        ws_premissas[f'C{row}'] = f'=B{row}*{CAMBIO_CELL}'
        ws_premissas[f'C{row}'].fill = calc_fill
        ws_premissas[f'C{row}'].border = thin_border
        ws_premissas[f'C{row}'].number_format = currency_format
        ws_premissas[f'D{row}'] = item[2]
        cloud_price_cells[item[0]] = f'B{row}'
        row += 1
    CLOUD_END_ROW = row - 1

    # Total cloud
    ws_premissas[f'A{row}'] = "TOTAL INFRA"
    ws_premissas[f'A{row}'].font = Font(bold=True)
    ws_premissas[f'B{row}'] = f'=SUM(B{CLOUD_START_ROW}:B{CLOUD_END_ROW})'
    ws_premissas[f'B{row}'].font = Font(bold=True)
    ws_premissas[f'B{row}'].fill = calc_fill
    ws_premissas[f'B{row}'].border = thin_border
    ws_premissas[f'B{row}'].number_format = currency_usd_format
    ws_premissas[f'C{row}'] = f'=B{row}*{CAMBIO_CELL}'
    ws_premissas[f'C{row}'].font = Font(bold=True)
    ws_premissas[f'C{row}'].fill = calc_fill
    ws_premissas[f'C{row}'].border = thin_border
    ws_premissas[f'C{row}'].number_format = currency_format
    TOTAL_INFRA_USD_CELL = f'B{row}'
    TOTAL_INFRA_BRL_CELL = f'C{row}'
    row += 2

    # ----- SEÇÃO: OPERAÇÕES POR WORKFLOW -----
    ws_premissas[f'A{row}'] = "OPERAÇÕES POR WORKFLOW"
    ws_premissas[f'A{row}'].font = header_font
    ws_premissas[f'A{row}'].fill = header_fill
    ws_premissas.merge_cells(f'A{row}:D{row}')
    row += 1

    ws_premissas[f'A{row}'] = "Operação"
    ws_premissas[f'B{row}'] = "Quantidade"
    ws_premissas[f'C{row}'] = "Unidade"
    ws_premissas[f'D{row}'] = "Fonte"
    for col in ['A', 'B', 'C', 'D']:
        ws_premissas[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws_premissas[f'{col}{row}'].fill = subheader_fill
    row += 1

    operations = [
        ("Temporal Actions", 9, "actions/wf", "Código: validation_workflow.go"),
        ("MongoDB Operations", 6, "ops/wf", "Código: activities.go (audit writes)"),
        ("Valkey Operations", 12, "ops/wf", "Código: locks + cache"),
    ]

    op_cells = {}
    for item in operations:
        ws_premissas[f'A{row}'] = item[0]
        ws_premissas[f'B{row}'] = item[1]
        ws_premissas[f'B{row}'].fill = input_fill
        ws_premissas[f'B{row}'].border = thin_border
        ws_premissas[f'C{row}'] = item[2]
        ws_premissas[f'D{row}'] = item[3]
        op_cells[item[0]] = f'B{row}'
        row += 1

    TEMPORAL_OPS_CELL = op_cells["Temporal Actions"]
    MONGO_OPS_CELL = op_cells["MongoDB Operations"]
    VALKEY_OPS_CELL = op_cells["Valkey Operations"]
    row += 1

    # ----- SEÇÃO: CUSTOS POR OPERAÇÃO -----
    ws_premissas[f'A{row}'] = "CUSTOS POR OPERAÇÃO (USD)"
    ws_premissas[f'A{row}'].font = header_font
    ws_premissas[f'A{row}'].fill = header_fill
    ws_premissas.merge_cells(f'A{row}:D{row}')
    row += 1

    ws_premissas[f'A{row}'] = "Operação"
    ws_premissas[f'B{row}'] = "Custo USD"
    ws_premissas[f'C{row}'] = "Unidade"
    ws_premissas[f'D{row}'] = "Fonte"
    for col in ['A', 'B', 'C', 'D']:
        ws_premissas[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws_premissas[f'{col}{row}'].fill = subheader_fill
    row += 1

    costs = [
        ("Temporal (por action)", 0.000025, "USD/action", "Temporal Cloud Pricing"),
        ("MongoDB (por operação)", 0.000001, "USD/op", "Estimativa Atlas"),
        ("Valkey (por operação)", 0.0000001, "USD/op", "Estimativa ElastiCache"),
    ]

    cost_cells = {}
    for item in costs:
        ws_premissas[f'A{row}'] = item[0]
        ws_premissas[f'B{row}'] = item[1]
        ws_premissas[f'B{row}'].fill = input_fill
        ws_premissas[f'B{row}'].border = thin_border
        ws_premissas[f'B{row}'].number_format = decimal_format
        ws_premissas[f'C{row}'] = item[2]
        ws_premissas[f'D{row}'] = item[3]
        cost_cells[item[0]] = f'B{row}'
        row += 1

    TEMPORAL_COST_CELL = cost_cells["Temporal (por action)"]
    MONGO_COST_CELL = cost_cells["MongoDB (por operação)"]
    VALKEY_COST_CELL = cost_cells["Valkey (por operação)"]
    row += 1

    # ----- SEÇÃO: CENÁRIO DE VOLUME -----
    ws_premissas[f'A{row}'] = "CENÁRIO DE VOLUME"
    ws_premissas[f'A{row}'].font = header_font
    ws_premissas[f'A{row}'].fill = header_fill
    ws_premissas.merge_cells(f'A{row}:D{row}')
    row += 1

    ws_premissas[f'A{row}'] = "Workflows por mês (cliente típico)"
    ws_premissas[f'B{row}'] = 25000
    ws_premissas[f'B{row}'].fill = input_fill
    ws_premissas[f'B{row}'].border = thin_border
    ws_premissas[f'B{row}'].number_format = number_format
    ws_premissas[f'D{row}'] = "← Volume médio esperado por cliente"
    VOLUME_CELL = f'B{row}'
    row += 1

    ws_premissas[f'A{row}'] = "% Infra compartilhada (best case)"
    ws_premissas[f'B{row}'] = 0.25
    ws_premissas[f'B{row}'].fill = input_fill
    ws_premissas[f'B{row}'].border = thin_border
    ws_premissas[f'B{row}'].number_format = percent_format
    ws_premissas[f'D{row}'] = "← Flowker usa X% do cluster compartilhado"
    SHARE_CELL = f'B{row}'
    row += 2

    # ----- SEÇÃO: PRICING TIERS -----
    ws_premissas[f'A{row}'] = "PRICING TIERS"
    ws_premissas[f'A{row}'].font = header_font
    ws_premissas[f'A{row}'].fill = header_fill
    ws_premissas.merge_cells(f'A{row}:D{row}')
    row += 1

    ws_premissas[f'A{row}'] = "Tier"
    ws_premissas[f'B{row}'] = "Base (BRL)"
    ws_premissas[f'C{row}'] = "Por WF (BRL)"
    ws_premissas[f'D{row}'] = "Limite WF/mês"
    for col in ['A', 'B', 'C', 'D']:
        ws_premissas[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws_premissas[f'{col}{row}'].fill = subheader_fill
    row += 1

    tiers = [
        ("Starter", 0, 0, 1000),
        ("Growth", 2990, 0.08, 25000),
        ("Scale", 9990, 0.04, 200000),
        ("Enterprise", 19990, 0.02, 1000000),
    ]

    tier_cells = {}
    TIER_START_ROW = row
    for tier in tiers:
        ws_premissas[f'A{row}'] = tier[0]
        ws_premissas[f'B{row}'] = tier[1]
        ws_premissas[f'B{row}'].fill = input_fill
        ws_premissas[f'B{row}'].border = thin_border
        ws_premissas[f'B{row}'].number_format = currency_format
        ws_premissas[f'C{row}'] = tier[2]
        ws_premissas[f'C{row}'].fill = input_fill
        ws_premissas[f'C{row}'].border = thin_border
        ws_premissas[f'C{row}'].number_format = 'R$ #,##0.00'
        ws_premissas[f'D{row}'] = tier[3]
        ws_premissas[f'D{row}'].fill = input_fill
        ws_premissas[f'D{row}'].border = thin_border
        ws_premissas[f'D{row}'].number_format = number_format
        tier_cells[tier[0]] = {'base': f'B{row}', 'per_wf': f'C{row}', 'limit': f'D{row}'}
        row += 1
    TIER_END_ROW = row - 1

    # Store important cell references for other sheets
    row += 2
    ws_premissas[f'A{row}'] = "REFERÊNCIAS INTERNAS (não editar)"
    ws_premissas[f'A{row}'].font = Font(bold=True, size=10, color="808080")
    row += 1
    ws_premissas[f'A{row}'] = "Custo Variável/WF (USD)"
    ws_premissas[f'B{row}'] = f'=({TEMPORAL_OPS_CELL}*{TEMPORAL_COST_CELL})+({MONGO_OPS_CELL}*{MONGO_COST_CELL})+({VALKEY_OPS_CELL}*{VALKEY_COST_CELL})'
    ws_premissas[f'B{row}'].fill = calc_fill
    ws_premissas[f'B{row}'].border = thin_border
    ws_premissas[f'B{row}'].number_format = decimal_format
    CUSTO_VAR_USD_CELL = f'B{row}'
    row += 1

    ws_premissas[f'A{row}'] = "Custo Variável/WF (BRL)"
    ws_premissas[f'B{row}'] = f'={CUSTO_VAR_USD_CELL}*{CAMBIO_CELL}'
    ws_premissas[f'B{row}'].fill = calc_fill
    ws_premissas[f'B{row}'].border = thin_border
    ws_premissas[f'B{row}'].number_format = decimal_format
    CUSTO_VAR_BRL_CELL = f'B{row}'

    # =========================================================================
    # ABA 2: CUSTOS
    # =========================================================================
    ws_custos = wb.create_sheet("CUSTOS")

    ws_custos.column_dimensions['A'].width = 35
    ws_custos.column_dimensions['B'].width = 18
    ws_custos.column_dimensions['C'].width = 18
    ws_custos.column_dimensions['D'].width = 18
    ws_custos.column_dimensions['E'].width = 15

    row = 1
    ws_custos['A1'] = "ANÁLISE DE CUSTOS - FLOWKER"
    ws_custos['A1'].font = Font(bold=True, size=16)
    ws_custos.merge_cells('A1:E1')
    row = 3

    # ----- CUSTOS FIXOS -----
    ws_custos[f'A{row}'] = "CUSTOS FIXOS MENSAIS (Infraestrutura Standalone)"
    ws_custos[f'A{row}'].font = header_font
    ws_custos[f'A{row}'].fill = header_fill
    ws_custos.merge_cells(f'A{row}:E{row}')
    row += 1

    ws_custos[f'A{row}'] = "Componente"
    ws_custos[f'B{row}'] = "Custo USD"
    ws_custos[f'C{row}'] = "Custo BRL"
    ws_custos[f'D{row}'] = "% do Total"
    ws_custos[f'E{row}'] = "Confiança"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_custos[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws_custos[f'{col}{row}'].fill = subheader_fill
    row += 1

    components = [
        ("PostgreSQL (RDS)", "PREMISSAS!B9"),
        ("MongoDB (Atlas)", "PREMISSAS!B10"),
        ("Valkey (ElastiCache)", "PREMISSAS!B11"),
        ("Temporal Cloud", "PREMISSAS!B12"),
        ("RabbitMQ (Amazon MQ)", "PREMISSAS!B13"),
        ("Compute (EKS)", "PREMISSAS!B14"),
    ]

    CUSTO_FIXO_START = row
    for comp in components:
        ws_custos[f'A{row}'] = comp[0]
        ws_custos[f'B{row}'] = f'={comp[1]}'
        ws_custos[f'B{row}'].fill = calc_fill
        ws_custos[f'B{row}'].border = thin_border
        ws_custos[f'B{row}'].number_format = currency_usd_format
        ws_custos[f'C{row}'] = f'=B{row}*PREMISSAS!B5'
        ws_custos[f'C{row}'].fill = calc_fill
        ws_custos[f'C{row}'].border = thin_border
        ws_custos[f'C{row}'].number_format = currency_format
        ws_custos[f'D{row}'] = f'=B{row}/B{CUSTO_FIXO_START + 6}'  # Reference to total
        ws_custos[f'D{row}'].fill = calc_fill
        ws_custos[f'D{row}'].border = thin_border
        ws_custos[f'D{row}'].number_format = percent_format
        ws_custos[f'E{row}'] = "🟢 Alto"
        row += 1
    CUSTO_FIXO_END = row - 1

    # Total
    ws_custos[f'A{row}'] = "TOTAL CUSTO FIXO"
    ws_custos[f'A{row}'].font = Font(bold=True)
    ws_custos[f'B{row}'] = f'=SUM(B{CUSTO_FIXO_START}:B{CUSTO_FIXO_END})'
    ws_custos[f'B{row}'].font = Font(bold=True)
    ws_custos[f'B{row}'].fill = calc_fill
    ws_custos[f'B{row}'].border = thin_border
    ws_custos[f'B{row}'].number_format = currency_usd_format
    ws_custos[f'C{row}'] = f'=B{row}*PREMISSAS!B5'
    ws_custos[f'C{row}'].font = Font(bold=True)
    ws_custos[f'C{row}'].fill = calc_fill
    ws_custos[f'C{row}'].border = thin_border
    ws_custos[f'C{row}'].number_format = currency_format
    ws_custos[f'D{row}'] = '=100%'
    ws_custos[f'D{row}'].fill = calc_fill
    ws_custos[f'D{row}'].border = thin_border
    ws_custos[f'D{row}'].number_format = percent_format
    TOTAL_FIXO_ROW = row
    row += 2

    # ----- CUSTOS VARIÁVEIS -----
    ws_custos[f'A{row}'] = "CUSTOS VARIÁVEIS (por Workflow)"
    ws_custos[f'A{row}'].font = header_font
    ws_custos[f'A{row}'].fill = header_fill
    ws_custos.merge_cells(f'A{row}:E{row}')
    row += 1

    ws_custos[f'A{row}'] = "Componente"
    ws_custos[f'B{row}'] = "Ops/WF"
    ws_custos[f'C{row}'] = "Custo/Op (USD)"
    ws_custos[f'D{row}'] = "Custo/WF (USD)"
    ws_custos[f'E{row}'] = "Confiança"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_custos[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws_custos[f'{col}{row}'].fill = subheader_fill
    row += 1

    var_costs = [
        ("Temporal Actions", "PREMISSAS!B19", "PREMISSAS!B26"),
        ("MongoDB Operations", "PREMISSAS!B20", "PREMISSAS!B27"),
        ("Valkey Operations", "PREMISSAS!B21", "PREMISSAS!B28"),
    ]

    CUSTO_VAR_START = row
    for vc in var_costs:
        ws_custos[f'A{row}'] = vc[0]
        ws_custos[f'B{row}'] = f'={vc[1]}'
        ws_custos[f'B{row}'].fill = calc_fill
        ws_custos[f'B{row}'].border = thin_border
        ws_custos[f'C{row}'] = f'={vc[2]}'
        ws_custos[f'C{row}'].fill = calc_fill
        ws_custos[f'C{row}'].border = thin_border
        ws_custos[f'C{row}'].number_format = decimal_format
        ws_custos[f'D{row}'] = f'=B{row}*C{row}'
        ws_custos[f'D{row}'].fill = calc_fill
        ws_custos[f'D{row}'].border = thin_border
        ws_custos[f'D{row}'].number_format = decimal_format
        ws_custos[f'E{row}'] = "🟠 Baixo-Médio"
        row += 1
    CUSTO_VAR_END = row - 1

    # Total variável
    ws_custos[f'A{row}'] = "TOTAL CUSTO VARIÁVEL/WF (USD)"
    ws_custos[f'A{row}'].font = Font(bold=True)
    ws_custos[f'D{row}'] = f'=SUM(D{CUSTO_VAR_START}:D{CUSTO_VAR_END})'
    ws_custos[f'D{row}'].font = Font(bold=True)
    ws_custos[f'D{row}'].fill = calc_fill
    ws_custos[f'D{row}'].border = thin_border
    ws_custos[f'D{row}'].number_format = decimal_format
    TOTAL_VAR_USD_ROW = row
    row += 1

    ws_custos[f'A{row}'] = "TOTAL CUSTO VARIÁVEL/WF (BRL)"
    ws_custos[f'A{row}'].font = Font(bold=True)
    ws_custos[f'D{row}'] = f'=D{TOTAL_VAR_USD_ROW}*PREMISSAS!B5'
    ws_custos[f'D{row}'].font = Font(bold=True)
    ws_custos[f'D{row}'].fill = calc_fill
    ws_custos[f'D{row}'].border = thin_border
    ws_custos[f'D{row}'].number_format = decimal_format
    TOTAL_VAR_BRL_ROW = row

    # =========================================================================
    # ABA 3: BREAK-EVEN
    # =========================================================================
    ws_breakeven = wb.create_sheet("BREAK_EVEN")

    ws_breakeven.column_dimensions['A'].width = 20
    ws_breakeven.column_dimensions['B'].width = 18
    ws_breakeven.column_dimensions['C'].width = 18
    ws_breakeven.column_dimensions['D'].width = 20
    ws_breakeven.column_dimensions['E'].width = 18

    row = 1
    ws_breakeven['A1'] = "ANÁLISE DE BREAK-EVEN"
    ws_breakeven['A1'].font = Font(bold=True, size=16)
    ws_breakeven.merge_cells('A1:E1')

    ws_breakeven['A2'] = "Fórmula: Break-even = Custo Fixo / (Preço/WF - Custo Variável/WF)"
    ws_breakeven['A2'].font = Font(italic=True)
    row = 4

    # References
    ws_breakeven[f'A{row}'] = "Custo Fixo Mensal (BRL):"
    ws_breakeven[f'B{row}'] = f'=CUSTOS!C{TOTAL_FIXO_ROW}'
    ws_breakeven[f'B{row}'].fill = calc_fill
    ws_breakeven[f'B{row}'].number_format = currency_format
    row += 1

    ws_breakeven[f'A{row}'] = "Custo Variável/WF (BRL):"
    ws_breakeven[f'B{row}'] = f'=CUSTOS!D{TOTAL_VAR_BRL_ROW}'
    ws_breakeven[f'B{row}'].fill = calc_fill
    ws_breakeven[f'B{row}'].number_format = decimal_format
    CUSTO_VAR_REF = f'B{row}'
    row += 2

    # Break-even table
    ws_breakeven[f'A{row}'] = "BREAK-EVEN POR PREÇO"
    ws_breakeven[f'A{row}'].font = header_font
    ws_breakeven[f'A{row}'].fill = header_fill
    ws_breakeven.merge_cells(f'A{row}:E{row}')
    row += 1

    ws_breakeven[f'A{row}'] = "Preço/WF (BRL)"
    ws_breakeven[f'B{row}'] = "Margem/WF (BRL)"
    ws_breakeven[f'C{row}'] = "Break-even (WF)"
    ws_breakeven[f'D{row}'] = "Clientes Equiv.*"
    ws_breakeven[f'E{row}'] = "Viabilidade"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_breakeven[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws_breakeven[f'{col}{row}'].fill = subheader_fill
    row += 1

    prices = [0.05, 0.08, 0.10, 0.15, 0.20]
    BE_START = row
    for price in prices:
        ws_breakeven[f'A{row}'] = price
        ws_breakeven[f'A{row}'].fill = input_fill
        ws_breakeven[f'A{row}'].border = thin_border
        ws_breakeven[f'A{row}'].number_format = 'R$ #,##0.00'
        # Margem = Preço - Custo Variável
        ws_breakeven[f'B{row}'] = f'=A{row}-{CUSTO_VAR_REF}'
        ws_breakeven[f'B{row}'].fill = calc_fill
        ws_breakeven[f'B{row}'].border = thin_border
        ws_breakeven[f'B{row}'].number_format = 'R$ #,##0.0000'
        # Break-even = Custo Fixo / Margem
        ws_breakeven[f'C{row}'] = f'=IF(B{row}>0,B4/B{row},"N/A")'
        ws_breakeven[f'C{row}'].fill = calc_fill
        ws_breakeven[f'C{row}'].border = thin_border
        ws_breakeven[f'C{row}'].number_format = number_format
        # Clientes equiv = Break-even / 25000
        ws_breakeven[f'D{row}'] = f'=IF(ISNUMBER(C{row}),C{row}/PREMISSAS!B31,"N/A")'
        ws_breakeven[f'D{row}'].fill = calc_fill
        ws_breakeven[f'D{row}'].border = thin_border
        ws_breakeven[f'D{row}'].number_format = '#,##0.0'
        # Viabilidade
        ws_breakeven[f'E{row}'] = f'=IF(D{row}<=1,"✅ Excelente",IF(D{row}<=2,"✅ Bom",IF(D{row}<=4,"⚠️ Aceitável","❌ Arriscado")))'
        ws_breakeven[f'E{row}'].fill = calc_fill
        ws_breakeven[f'E{row}'].border = thin_border
        row += 1

    row += 1
    ws_breakeven[f'A{row}'] = "*Assumindo cliente médio com volume da aba PREMISSAS"
    ws_breakeven[f'A{row}'].font = Font(italic=True, size=9)

    # =========================================================================
    # ABA 4: CENÁRIOS
    # =========================================================================
    ws_cenarios = wb.create_sheet("CENARIOS")

    ws_cenarios.column_dimensions['A'].width = 30
    ws_cenarios.column_dimensions['B'].width = 18
    ws_cenarios.column_dimensions['C'].width = 18
    ws_cenarios.column_dimensions['D'].width = 18

    row = 1
    ws_cenarios['A1'] = "CENÁRIOS DE MARGEM"
    ws_cenarios['A1'].font = Font(bold=True, size=16)
    ws_cenarios.merge_cells('A1:D1')
    row = 3

    # Cenário 1: Standalone
    ws_cenarios[f'A{row}'] = "CENÁRIO 1: STANDALONE (Worst Case)"
    ws_cenarios[f'A{row}'].font = header_font
    ws_cenarios[f'A{row}'].fill = header_fill
    ws_cenarios.merge_cells(f'A{row}:D{row}')
    row += 1

    ws_cenarios[f'A{row}'] = "Premissa: 1 cliente, infra 100% dedicada"
    ws_cenarios[f'A{row}'].font = Font(italic=True)
    row += 2

    ws_cenarios[f'A{row}'] = "Workflows/mês:"
    ws_cenarios[f'B{row}'] = '=PREMISSAS!B31'
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = number_format
    STANDALONE_VOL = f'B{row}'
    row += 1

    ws_cenarios[f'A{row}'] = "Tier:"
    ws_cenarios[f'B{row}'] = "Growth"
    row += 1

    ws_cenarios[f'A{row}'] = "Receita Base:"
    ws_cenarios[f'B{row}'] = '=PREMISSAS!B37'
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = currency_format
    STANDALONE_BASE = f'B{row}'
    row += 1

    ws_cenarios[f'A{row}'] = "Receita Variável:"
    ws_cenarios[f'B{row}'] = f'={STANDALONE_VOL}*PREMISSAS!C37'
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = currency_format
    STANDALONE_VAR_REC = f'B{row}'
    row += 1

    ws_cenarios[f'A{row}'] = "RECEITA TOTAL:"
    ws_cenarios[f'A{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'] = f'={STANDALONE_BASE}+{STANDALONE_VAR_REC}'
    ws_cenarios[f'B{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = currency_format
    STANDALONE_RECEITA = f'B{row}'
    row += 2

    ws_cenarios[f'A{row}'] = "Custo Fixo (100% infra):"
    ws_cenarios[f'B{row}'] = f'=CUSTOS!C{TOTAL_FIXO_ROW}'
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = currency_format
    STANDALONE_FIXO = f'B{row}'
    row += 1

    ws_cenarios[f'A{row}'] = "Custo Variável:"
    ws_cenarios[f'B{row}'] = f'={STANDALONE_VOL}*CUSTOS!D{TOTAL_VAR_BRL_ROW}'
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = currency_format
    STANDALONE_VAR_CUSTO = f'B{row}'
    row += 1

    ws_cenarios[f'A{row}'] = "CUSTO TOTAL:"
    ws_cenarios[f'A{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'] = f'={STANDALONE_FIXO}+{STANDALONE_VAR_CUSTO}'
    ws_cenarios[f'B{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = currency_format
    STANDALONE_CUSTO = f'B{row}'
    row += 2

    ws_cenarios[f'A{row}'] = "MARGEM BRUTA (R$):"
    ws_cenarios[f'A{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'] = f'={STANDALONE_RECEITA}-{STANDALONE_CUSTO}'
    ws_cenarios[f'B{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = currency_format
    STANDALONE_MARGEM = f'B{row}'
    row += 1

    ws_cenarios[f'A{row}'] = "MARGEM BRUTA (%):"
    ws_cenarios[f'A{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'] = f'={STANDALONE_MARGEM}/{STANDALONE_RECEITA}'
    ws_cenarios[f'B{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = percent_format
    row += 3

    # Cenário 2: Compartilhado
    ws_cenarios[f'A{row}'] = "CENÁRIO 2: COMPARTILHADO (Best Case)"
    ws_cenarios[f'A{row}'].font = header_font
    ws_cenarios[f'A{row}'].fill = header_fill
    ws_cenarios.merge_cells(f'A{row}:D{row}')
    row += 1

    ws_cenarios[f'A{row}'] = "Premissa: Múltiplos clientes, infra compartilhada"
    ws_cenarios[f'A{row}'].font = Font(italic=True)
    row += 2

    ws_cenarios[f'A{row}'] = "% da Infra usada:"
    ws_cenarios[f'B{row}'] = '=PREMISSAS!B32'
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = percent_format
    SHARE_REF = f'B{row}'
    row += 1

    ws_cenarios[f'A{row}'] = "Workflows/mês:"
    ws_cenarios[f'B{row}'] = '=PREMISSAS!B31'
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = number_format
    SHARED_VOL = f'B{row}'
    row += 1

    ws_cenarios[f'A{row}'] = "RECEITA TOTAL:"
    ws_cenarios[f'A{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'] = f'={STANDALONE_RECEITA}'
    ws_cenarios[f'B{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = currency_format
    SHARED_RECEITA = f'B{row}'
    row += 2

    ws_cenarios[f'A{row}'] = "Custo Fixo (% compartilhado):"
    ws_cenarios[f'B{row}'] = f'=CUSTOS!C{TOTAL_FIXO_ROW}*{SHARE_REF}'
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = currency_format
    SHARED_FIXO = f'B{row}'
    row += 1

    ws_cenarios[f'A{row}'] = "Custo Variável:"
    ws_cenarios[f'B{row}'] = f'={SHARED_VOL}*CUSTOS!D{TOTAL_VAR_BRL_ROW}'
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = currency_format
    SHARED_VAR_CUSTO = f'B{row}'
    row += 1

    ws_cenarios[f'A{row}'] = "CUSTO TOTAL:"
    ws_cenarios[f'A{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'] = f'={SHARED_FIXO}+{SHARED_VAR_CUSTO}'
    ws_cenarios[f'B{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = currency_format
    SHARED_CUSTO = f'B{row}'
    row += 2

    ws_cenarios[f'A{row}'] = "MARGEM BRUTA (R$):"
    ws_cenarios[f'A{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'] = f'={SHARED_RECEITA}-{SHARED_CUSTO}'
    ws_cenarios[f'B{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = currency_format
    SHARED_MARGEM = f'B{row}'
    row += 1

    ws_cenarios[f'A{row}'] = "MARGEM BRUTA (%):"
    ws_cenarios[f'A{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'] = f'={SHARED_MARGEM}/{SHARED_RECEITA}'
    ws_cenarios[f'B{row}'].font = Font(bold=True)
    ws_cenarios[f'B{row}'].fill = calc_fill
    ws_cenarios[f'B{row}'].number_format = percent_format

    # =========================================================================
    # ABA 5: PRICING TIERS
    # =========================================================================
    ws_pricing = wb.create_sheet("PRICING")

    ws_pricing.column_dimensions['A'].width = 15
    ws_pricing.column_dimensions['B'].width = 15
    ws_pricing.column_dimensions['C'].width = 15
    ws_pricing.column_dimensions['D'].width = 15
    ws_pricing.column_dimensions['E'].width = 18
    ws_pricing.column_dimensions['F'].width = 18
    ws_pricing.column_dimensions['G'].width = 18
    ws_pricing.column_dimensions['H'].width = 15

    row = 1
    ws_pricing['A1'] = "MODELO DE PRICING - FLOWKER"
    ws_pricing['A1'].font = Font(bold=True, size=16)
    ws_pricing.merge_cells('A1:H1')

    ws_pricing['A2'] = "Simulação com base nas premissas (ajuste na aba PREMISSAS)"
    ws_pricing['A2'].font = Font(italic=True)
    row = 4

    ws_pricing[f'A{row}'] = "SIMULAÇÃO POR TIER"
    ws_pricing[f'A{row}'].font = header_font
    ws_pricing[f'A{row}'].fill = header_fill
    ws_pricing.merge_cells(f'A{row}:H{row}')
    row += 1

    headers = ["Tier", "Base (BRL)", "Por WF", "Volume", "Receita", "Custo*", "Margem R$", "Margem %"]
    for i, h in enumerate(headers):
        col = get_column_letter(i+1)
        ws_pricing[f'{col}{row}'] = h
        ws_pricing[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws_pricing[f'{col}{row}'].fill = subheader_fill
    row += 1

    # Starter
    ws_pricing[f'A{row}'] = "Starter"
    ws_pricing[f'B{row}'] = '=PREMISSAS!B36'
    ws_pricing[f'B{row}'].number_format = currency_format
    ws_pricing[f'C{row}'] = '=PREMISSAS!C36'
    ws_pricing[f'C{row}'].number_format = 'R$ #,##0.00'
    ws_pricing[f'D{row}'] = '=PREMISSAS!D36'
    ws_pricing[f'D{row}'].number_format = number_format
    ws_pricing[f'E{row}'] = f'=B{row}+(D{row}*C{row})'
    ws_pricing[f'E{row}'].number_format = currency_format
    ws_pricing[f'E{row}'].fill = calc_fill
    ws_pricing[f'F{row}'] = f'=CUSTOS!C{TOTAL_FIXO_ROW}+(D{row}*CUSTOS!D{TOTAL_VAR_BRL_ROW})'
    ws_pricing[f'F{row}'].number_format = currency_format
    ws_pricing[f'F{row}'].fill = calc_fill
    ws_pricing[f'G{row}'] = f'=E{row}-F{row}'
    ws_pricing[f'G{row}'].number_format = currency_format
    ws_pricing[f'G{row}'].fill = calc_fill
    ws_pricing[f'H{row}'] = f'=IF(E{row}>0,G{row}/E{row},0)'
    ws_pricing[f'H{row}'].number_format = percent_format
    ws_pricing[f'H{row}'].fill = calc_fill
    row += 1

    # Growth
    ws_pricing[f'A{row}'] = "Growth"
    ws_pricing[f'B{row}'] = '=PREMISSAS!B37'
    ws_pricing[f'B{row}'].number_format = currency_format
    ws_pricing[f'C{row}'] = '=PREMISSAS!C37'
    ws_pricing[f'C{row}'].number_format = 'R$ #,##0.00'
    ws_pricing[f'D{row}'] = '=PREMISSAS!D37'
    ws_pricing[f'D{row}'].number_format = number_format
    ws_pricing[f'E{row}'] = f'=B{row}+(D{row}*C{row})'
    ws_pricing[f'E{row}'].number_format = currency_format
    ws_pricing[f'E{row}'].fill = calc_fill
    ws_pricing[f'F{row}'] = f'=CUSTOS!C{TOTAL_FIXO_ROW}+(D{row}*CUSTOS!D{TOTAL_VAR_BRL_ROW})'
    ws_pricing[f'F{row}'].number_format = currency_format
    ws_pricing[f'F{row}'].fill = calc_fill
    ws_pricing[f'G{row}'] = f'=E{row}-F{row}'
    ws_pricing[f'G{row}'].number_format = currency_format
    ws_pricing[f'G{row}'].fill = calc_fill
    ws_pricing[f'H{row}'] = f'=IF(E{row}>0,G{row}/E{row},0)'
    ws_pricing[f'H{row}'].number_format = percent_format
    ws_pricing[f'H{row}'].fill = calc_fill
    row += 1

    # Scale
    ws_pricing[f'A{row}'] = "Scale"
    ws_pricing[f'B{row}'] = '=PREMISSAS!B38'
    ws_pricing[f'B{row}'].number_format = currency_format
    ws_pricing[f'C{row}'] = '=PREMISSAS!C38'
    ws_pricing[f'C{row}'].number_format = 'R$ #,##0.00'
    ws_pricing[f'D{row}'] = '=PREMISSAS!D38'
    ws_pricing[f'D{row}'].number_format = number_format
    ws_pricing[f'E{row}'] = f'=B{row}+(D{row}*C{row})'
    ws_pricing[f'E{row}'].number_format = currency_format
    ws_pricing[f'E{row}'].fill = calc_fill
    ws_pricing[f'F{row}'] = f'=CUSTOS!C{TOTAL_FIXO_ROW}+(D{row}*CUSTOS!D{TOTAL_VAR_BRL_ROW})'
    ws_pricing[f'F{row}'].number_format = currency_format
    ws_pricing[f'F{row}'].fill = calc_fill
    ws_pricing[f'G{row}'] = f'=E{row}-F{row}'
    ws_pricing[f'G{row}'].number_format = currency_format
    ws_pricing[f'G{row}'].fill = calc_fill
    ws_pricing[f'H{row}'] = f'=IF(E{row}>0,G{row}/E{row},0)'
    ws_pricing[f'H{row}'].number_format = percent_format
    ws_pricing[f'H{row}'].fill = calc_fill
    row += 1

    # Enterprise
    ws_pricing[f'A{row}'] = "Enterprise"
    ws_pricing[f'B{row}'] = '=PREMISSAS!B39'
    ws_pricing[f'B{row}'].number_format = currency_format
    ws_pricing[f'C{row}'] = '=PREMISSAS!C39'
    ws_pricing[f'C{row}'].number_format = 'R$ #,##0.00'
    ws_pricing[f'D{row}'] = '=PREMISSAS!D39'
    ws_pricing[f'D{row}'].number_format = number_format
    ws_pricing[f'E{row}'] = f'=B{row}+(D{row}*C{row})'
    ws_pricing[f'E{row}'].number_format = currency_format
    ws_pricing[f'E{row}'].fill = calc_fill
    ws_pricing[f'F{row}'] = f'=CUSTOS!C{TOTAL_FIXO_ROW}+(D{row}*CUSTOS!D{TOTAL_VAR_BRL_ROW})'
    ws_pricing[f'F{row}'].number_format = currency_format
    ws_pricing[f'F{row}'].fill = calc_fill
    ws_pricing[f'G{row}'] = f'=E{row}-F{row}'
    ws_pricing[f'G{row}'].number_format = currency_format
    ws_pricing[f'G{row}'].fill = calc_fill
    ws_pricing[f'H{row}'] = f'=IF(E{row}>0,G{row}/E{row},0)'
    ws_pricing[f'H{row}'].number_format = percent_format
    ws_pricing[f'H{row}'].fill = calc_fill
    row += 2

    ws_pricing[f'A{row}'] = "*Custo assume infra STANDALONE (worst case)"
    ws_pricing[f'A{row}'].font = Font(italic=True, size=9)
    row += 3

    # Simulador customizado
    ws_pricing[f'A{row}'] = "SIMULADOR CUSTOMIZADO"
    ws_pricing[f'A{row}'].font = header_font
    ws_pricing[f'A{row}'].fill = header_fill
    ws_pricing.merge_cells(f'A{row}:H{row}')
    row += 1

    ws_pricing[f'A{row}'] = "Insira valores:"
    ws_pricing[f'A{row}'].font = Font(bold=True)
    row += 1

    ws_pricing[f'A{row}'] = "Preço Base (BRL):"
    ws_pricing[f'B{row}'] = 2990
    ws_pricing[f'B{row}'].fill = input_fill
    ws_pricing[f'B{row}'].border = thin_border
    ws_pricing[f'B{row}'].number_format = currency_format
    CUSTOM_BASE = f'B{row}'
    row += 1

    ws_pricing[f'A{row}'] = "Preço/WF (BRL):"
    ws_pricing[f'B{row}'] = 0.08
    ws_pricing[f'B{row}'].fill = input_fill
    ws_pricing[f'B{row}'].border = thin_border
    ws_pricing[f'B{row}'].number_format = 'R$ #,##0.00'
    CUSTOM_PER_WF = f'B{row}'
    row += 1

    ws_pricing[f'A{row}'] = "Volume (WF/mês):"
    ws_pricing[f'B{row}'] = 25000
    ws_pricing[f'B{row}'].fill = input_fill
    ws_pricing[f'B{row}'].border = thin_border
    ws_pricing[f'B{row}'].number_format = number_format
    CUSTOM_VOL = f'B{row}'
    row += 2

    ws_pricing[f'A{row}'] = "Resultados:"
    ws_pricing[f'A{row}'].font = Font(bold=True)
    row += 1

    ws_pricing[f'A{row}'] = "Receita:"
    ws_pricing[f'B{row}'] = f'={CUSTOM_BASE}+({CUSTOM_VOL}*{CUSTOM_PER_WF})'
    ws_pricing[f'B{row}'].fill = calc_fill
    ws_pricing[f'B{row}'].number_format = currency_format
    CUSTOM_REC = f'B{row}'
    row += 1

    ws_pricing[f'A{row}'] = "Custo (Standalone):"
    ws_pricing[f'B{row}'] = f'=CUSTOS!C{TOTAL_FIXO_ROW}+({CUSTOM_VOL}*CUSTOS!D{TOTAL_VAR_BRL_ROW})'
    ws_pricing[f'B{row}'].fill = calc_fill
    ws_pricing[f'B{row}'].number_format = currency_format
    CUSTOM_CUSTO = f'B{row}'
    row += 1

    ws_pricing[f'A{row}'] = "Margem (R$):"
    ws_pricing[f'B{row}'] = f'={CUSTOM_REC}-{CUSTOM_CUSTO}'
    ws_pricing[f'B{row}'].fill = calc_fill
    ws_pricing[f'B{row}'].number_format = currency_format
    CUSTOM_MARGEM = f'B{row}'
    row += 1

    ws_pricing[f'A{row}'] = "Margem (%):"
    ws_pricing[f'B{row}'] = f'=IF({CUSTOM_REC}>0,{CUSTOM_MARGEM}/{CUSTOM_REC},0)'
    ws_pricing[f'B{row}'].fill = calc_fill
    ws_pricing[f'B{row}'].number_format = percent_format
    row += 1

    ws_pricing[f'A{row}'] = "Status:"
    ws_pricing[f'B{row}'] = f'=IF({CUSTOM_MARGEM}>0,"✅ Sustentável","❌ Prejuízo")'
    ws_pricing[f'B{row}'].fill = calc_fill

    # =========================================================================
    # ABA 6: CONFIANÇA
    # =========================================================================
    ws_confianca = wb.create_sheet("CONFIANCA")

    ws_confianca.column_dimensions['A'].width = 40
    ws_confianca.column_dimensions['B'].width = 20
    ws_confianca.column_dimensions['C'].width = 50

    row = 1
    ws_confianca['A1'] = "GRAUS DE CONFIANÇA"
    ws_confianca['A1'].font = Font(bold=True, size=16)
    ws_confianca.merge_cells('A1:C1')
    row = 3

    ws_confianca[f'A{row}'] = "Componente"
    ws_confianca[f'B{row}'] = "Confiança"
    ws_confianca[f'C{row}'] = "Justificativa"
    for col in ['A', 'B', 'C']:
        ws_confianca[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
        ws_confianca[f'{col}{row}'].fill = header_fill
    row += 1

    confianca_items = [
        ("Metodologia (standalone-first)", "🟢 Alto", "Princípio de negócio, não técnico"),
        ("Componentes de infraestrutura", "🟢 Alto", "Baseado em docker-compose real"),
        ("Preços de cloud (unitários)", "🟢 Alto", "Pricing público verificável AWS/MongoDB"),
        ("Sizing de produção", "🟠 Baixo-Médio", "Sem Terraform/load tests - estimar conservador"),
        ("Operações por workflow", "🟠 Baixo-Médio", "Baseado em código, não em métricas reais"),
        ("Custo variável por workflow", "🟠 Baixo-Médio", "Depende de validação com benchmark"),
        ("Modelo de pricing (estrutura)", "🟢 Alto", "Padrão de mercado (base + variável)"),
        ("Modelo de pricing (valores)", "🟡 Médio", "Depende de validação de custos"),
    ]

    for item in confianca_items:
        ws_confianca[f'A{row}'] = item[0]
        ws_confianca[f'B{row}'] = item[1]
        ws_confianca[f'C{row}'] = item[2]
        row += 1

    row += 2
    ws_confianca[f'A{row}'] = "AÇÕES PARA AUMENTAR CONFIANÇA"
    ws_confianca[f'A{row}'].font = Font(bold=True)
    row += 1

    acoes = [
        "1. Executar benchmark script com Docker rodando",
        "2. Instalar KubeCost no cluster K8s",
        "3. Load test com 25K-100K workflows",
        "4. Criar Terraform do Flowker",
        "5. Coletar 30+ dias de dados de billing",
    ]
    for acao in acoes:
        ws_confianca[f'A{row}'] = acao
        row += 1

    # Save
    output_path = "/tmp/pricing-flowker/Flowker_Pricing_Model.xlsx"
    wb.save(output_path)
    print(f"Planilha criada: {output_path}")
    return output_path

if __name__ == "__main__":
    create_pricing_spreadsheet()
