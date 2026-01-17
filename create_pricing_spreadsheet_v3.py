#!/usr/bin/env python3
"""
Flowker Pricing Analysis Spreadsheet - V3 (REVISADO 3x)
Cross-checked com PDF analise_pricing_flowker_v2.md

MAPA DE CÉLULAS DEFINITIVO:
==========================
Aba PREMISSAS:
- B5  = Taxa de câmbio (5.00)
- B9  = PostgreSQL ($108)
- B10 = MongoDB ($389)
- B11 = Valkey ($94)
- B12 = Temporal ($25)
- B13 = RabbitMQ ($180)
- B14 = Compute ($122)
- B15 = TOTAL FIXO USD ($918)
- C15 = TOTAL FIXO BRL (R$ 4.590)
- B19 = Temporal ops/wf (9)
- B20 = MongoDB ops/wf (6)
- B21 = Valkey ops/wf (12)
- B25 = Custo Temporal/action ($0.000025)
- B26 = Custo MongoDB/op ($0.000001)
- B27 = Custo Valkey/op ($0.0000001)
- B31 = Volume wf/mês (25000)
- B35 = Starter base (0)
- B36 = Growth base (2990)
- B37 = Scale base (9990)
- B41 = Custo variável/wf USD
- B42 = Custo variável/wf BRL

Aba BREAK_EVEN:
- B4 = Custo Fixo BRL (=PREMISSAS!C15)
- B5 = Custo Var/wf BRL (=PREMISSAS!B42)
- B6 = Volume típico (=PREMISSAS!B31)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

def create_pricing_spreadsheet():
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # =========================================================================
    # ABA: PREMISSAS
    # =========================================================================
    ws = wb.active
    ws.title = "PREMISSAS"

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 40

    # Row 1-2: Títulos
    ws['A1'] = "ANÁLISE DE PRICING - FLOWKER"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = "AMARELO = Input editável | VERDE = Calculado"
    ws['A2'].font = Font(italic=True, size=10)

    # Row 4-5: CÂMBIO
    ws['A4'] = "TAXA DE CÂMBIO"
    ws['A4'].font = header_font
    ws['A4'].fill = header_fill

    ws['A5'] = "USD para BRL"
    ws['B5'] = 5.00  # <<<< CÉLULA B5 = CÂMBIO
    ws['B5'].fill = input_fill
    ws['B5'].border = border
    ws['C5'] = "BRL/USD"

    # Row 7-15: CUSTOS DE INFRAESTRUTURA
    ws['A7'] = "CUSTOS FIXOS - INFRAESTRUTURA STANDALONE (USD/mês)"
    ws['A7'].font = header_font
    ws['A7'].fill = header_fill

    ws['A8'] = "Componente"
    ws['B8'] = "USD/mês"
    ws['C8'] = "BRL/mês"
    ws['D8'] = "Configuração"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}8'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}8'].fill = subheader_fill

    # Custos de cloud - VALORES DO PDF
    cloud_costs = [
        (9, "PostgreSQL (RDS db.r6g.large)", 108, "2 vCPU, 16GB RAM"),
        (10, "MongoDB (Atlas M30)", 389, "2 vCPU, 8GB RAM"),
        (11, "Valkey (ElastiCache cache.r6g.large)", 94, "2 vCPU, 13GB RAM"),
        (12, "Temporal Cloud (base)", 25, "Orquestração workflows"),
        (13, "RabbitMQ (Amazon MQ mq.m5.large)", 180, "2 vCPU, 8GB RAM"),
        (14, "Compute (EKS 2x c6g.large)", 122, "4 vCPU total, 8GB RAM"),
    ]

    for row, name, usd, config in cloud_costs:
        ws[f'A{row}'] = name
        ws[f'B{row}'] = usd
        ws[f'B{row}'].fill = input_fill
        ws[f'B{row}'].border = border
        ws[f'B{row}'].number_format = '$ #,##0.00'
        ws[f'C{row}'] = f'=B{row}*$B$5'  # Referência absoluta ao câmbio
        ws[f'C{row}'].fill = calc_fill
        ws[f'C{row}'].border = border
        ws[f'C{row}'].number_format = 'R$ #,##0.00'
        ws[f'D{row}'] = config

    # Total (row 15)
    ws['A15'] = "TOTAL CUSTO FIXO MENSAL"
    ws['A15'].font = Font(bold=True)
    ws['B15'] = '=SUM(B9:B14)'  # <<<< B15 = Total USD
    ws['B15'].font = Font(bold=True)
    ws['B15'].fill = calc_fill
    ws['B15'].border = border
    ws['B15'].number_format = '$ #,##0.00'
    ws['C15'] = '=B15*$B$5'  # <<<< C15 = Total BRL (R$ 4.590)
    ws['C15'].font = Font(bold=True)
    ws['C15'].fill = calc_fill
    ws['C15'].border = border
    ws['C15'].number_format = 'R$ #,##0.00'

    # Row 17-21: OPERAÇÕES POR WORKFLOW
    ws['A17'] = "OPERAÇÕES POR WORKFLOW (análise de código)"
    ws['A17'].font = header_font
    ws['A17'].fill = header_fill

    ws['A18'] = "Operação"
    ws['B18'] = "Qtd/WF"
    ws['C18'] = "Unidade"
    ws['D18'] = "Arquivo fonte"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}18'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}18'].fill = subheader_fill

    ops = [
        (19, "Temporal Actions", 9, "actions", "validation_workflow.go"),
        (20, "MongoDB Operations", 6, "ops", "activities.go (audit)"),
        (21, "Valkey Operations", 12, "ops", "activities.go (locks)"),
    ]

    for row, name, qty, unit, source in ops:
        ws[f'A{row}'] = name
        ws[f'B{row}'] = qty  # <<<< B19=9, B20=6, B21=12
        ws[f'B{row}'].fill = input_fill
        ws[f'B{row}'].border = border
        ws[f'C{row}'] = unit
        ws[f'D{row}'] = source

    # Row 23-27: CUSTOS POR OPERAÇÃO
    ws['A23'] = "CUSTOS POR OPERAÇÃO (USD)"
    ws['A23'].font = header_font
    ws['A23'].fill = header_fill

    ws['A24'] = "Operação"
    ws['B24'] = "USD/op"
    ws['C24'] = "BRL/op"
    ws['D24'] = "Fonte"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}24'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}24'].fill = subheader_fill

    costs = [
        (25, "Temporal (por action)", 0.000025, "Temporal Cloud Pricing"),
        (26, "MongoDB (por operação)", 0.000001, "Estimativa Atlas"),
        (27, "Valkey (por operação)", 0.0000001, "Estimativa ElastiCache"),
    ]

    for row, name, cost, source in costs:
        ws[f'A{row}'] = name
        ws[f'B{row}'] = cost  # <<<< B25=0.000025, B26=0.000001, B27=0.0000001
        ws[f'B{row}'].fill = input_fill
        ws[f'B{row}'].border = border
        ws[f'B{row}'].number_format = '$ #,##0.000000'
        ws[f'C{row}'] = f'=B{row}*$B$5'
        ws[f'C{row}'].fill = calc_fill
        ws[f'C{row}'].border = border
        ws[f'C{row}'].number_format = 'R$ #,##0.000000'
        ws[f'D{row}'] = source

    # Row 29-31: VOLUME
    ws['A29'] = "VOLUME DE WORKFLOWS"
    ws['A29'].font = header_font
    ws['A29'].fill = header_fill

    ws['A30'] = "Descrição"
    ws['B30'] = "Valor"
    ws['A30'].font = Font(bold=True, color="FFFFFF")
    ws['A30'].fill = subheader_fill
    ws['B30'].font = Font(bold=True, color="FFFFFF")
    ws['B30'].fill = subheader_fill

    ws['A31'] = "Workflows/mês (cliente típico Growth)"
    ws['B31'] = 25000  # <<<< B31 = VOLUME TÍPICO
    ws['B31'].fill = input_fill
    ws['B31'].border = border
    ws['B31'].number_format = '#,##0'

    # Row 33-38: PRICING TIERS
    ws['A33'] = "PRICING TIERS"
    ws['A33'].font = header_font
    ws['A33'].fill = header_fill

    ws['A34'] = "Tier"
    ws['B34'] = "Base (BRL)"
    ws['C34'] = "Por WF (BRL)"
    ws['D34'] = "Limite WF/mês"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}34'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}34'].fill = subheader_fill

    tiers = [
        (35, "Starter", 0, 0.00, 1000),
        (36, "Growth", 2990, 0.08, 25000),
        (37, "Scale", 9990, 0.04, 200000),
        (38, "Enterprise", 19990, 0.02, 1000000),
    ]

    for row, name, base, per_wf, limit in tiers:
        ws[f'A{row}'] = name
        ws[f'B{row}'] = base
        ws[f'B{row}'].fill = input_fill
        ws[f'B{row}'].border = border
        ws[f'B{row}'].number_format = 'R$ #,##0.00'
        ws[f'C{row}'] = per_wf
        ws[f'C{row}'].fill = input_fill
        ws[f'C{row}'].border = border
        ws[f'C{row}'].number_format = 'R$ #,##0.00'
        ws[f'D{row}'] = limit
        ws[f'D{row}'].fill = input_fill
        ws[f'D{row}'].border = border
        ws[f'D{row}'].number_format = '#,##0'

    # Row 40-42: CÁLCULOS DERIVADOS
    ws['A40'] = "CÁLCULOS DERIVADOS (não editar)"
    ws['A40'].font = header_font
    ws['A40'].fill = header_fill

    ws['A41'] = "Custo Variável/WF (USD)"
    # Fórmula: (9 * 0.000025) + (6 * 0.000001) + (12 * 0.0000001) = 0.0002322
    ws['B41'] = '=(B19*B25)+(B20*B26)+(B21*B27)'  # <<<< B41 = Custo var USD
    ws['B41'].fill = calc_fill
    ws['B41'].border = border
    ws['B41'].number_format = '$ #,##0.000000'

    ws['A42'] = "Custo Variável/WF (BRL)"
    # Fórmula: 0.0002322 * 5 = 0.001161
    ws['B42'] = '=B41*$B$5'  # <<<< B42 = Custo var BRL (R$ 0.00116)
    ws['B42'].fill = calc_fill
    ws['B42'].border = border
    ws['B42'].number_format = 'R$ #,##0.000000'

    # Verificação
    ws['A44'] = "VERIFICAÇÃO (deve bater com PDF)"
    ws['A44'].font = Font(bold=True, color="666666")
    ws['A45'] = "Total Fixo esperado: R$ 4.590,00"
    ws['A45'].font = Font(size=9, color="666666")
    ws['A46'] = "Custo Var esperado: R$ 0,00116"
    ws['A46'].font = Font(size=9, color="666666")

    # =========================================================================
    # ABA: BREAK_EVEN
    # =========================================================================
    ws2 = wb.create_sheet("BREAK_EVEN")

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 15

    ws2['A1'] = "ANÁLISE DE BREAK-EVEN (Standalone)"
    ws2['A1'].font = Font(bold=True, size=16)

    ws2['A2'] = "Break-even = Custo Fixo / (Preço/WF - Custo Var/WF)"
    ws2['A2'].font = Font(italic=True)

    # Referências (rows 4-6)
    ws2['A4'] = "Custo Fixo Mensal (BRL):"
    ws2['B4'] = '=PREMISSAS!C15'  # <<<< Referencia C15 (R$ 4.590)
    ws2['B4'].fill = calc_fill
    ws2['B4'].border = border
    ws2['B4'].number_format = 'R$ #,##0.00'

    ws2['A5'] = "Custo Variável/WF (BRL):"
    ws2['B5'] = '=PREMISSAS!B42'  # <<<< Referencia B42 (R$ 0.00116)
    ws2['B5'].fill = calc_fill
    ws2['B5'].border = border
    ws2['B5'].number_format = 'R$ #,##0.000000'

    ws2['A6'] = "Volume típico (WF/mês):"
    ws2['B6'] = '=PREMISSAS!B31'  # <<<< Referencia B31 (25000)
    ws2['B6'].fill = calc_fill
    ws2['B6'].border = border
    ws2['B6'].number_format = '#,##0'

    # Tabela break-even (row 8+)
    ws2['A8'] = "BREAK-EVEN POR PREÇO"
    ws2['A8'].font = header_font
    ws2['A8'].fill = header_fill

    ws2['A9'] = "Preço/WF"
    ws2['B9'] = "Margem/WF"
    ws2['C9'] = "Break-even"
    ws2['D9'] = "Clientes*"
    ws2['E9'] = "Status"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws2[f'{col}9'].font = Font(bold=True, color="FFFFFF")
        ws2[f'{col}9'].fill = subheader_fill

    # Preços para análise (do PDF)
    prices = [
        (10, 0.05),
        (11, 0.08),
        (12, 0.10),
        (13, 0.15),
        (14, 0.20),
    ]

    for row, price in prices:
        ws2[f'A{row}'] = price
        ws2[f'A{row}'].fill = input_fill
        ws2[f'A{row}'].border = border
        ws2[f'A{row}'].number_format = 'R$ #,##0.00'

        # Margem = Preço - Custo Var (B5)
        ws2[f'B{row}'] = f'=A{row}-$B$5'
        ws2[f'B{row}'].fill = calc_fill
        ws2[f'B{row}'].border = border
        ws2[f'B{row}'].number_format = 'R$ #,##0.0000'

        # Break-even = Custo Fixo (B4) / Margem
        ws2[f'C{row}'] = f'=IF(B{row}>0,$B$4/B{row},"N/A")'
        ws2[f'C{row}'].fill = calc_fill
        ws2[f'C{row}'].border = border
        ws2[f'C{row}'].number_format = '#,##0'

        # Clientes = Break-even / Volume (B6)
        ws2[f'D{row}'] = f'=IF(ISNUMBER(C{row}),C{row}/$B$6,"N/A")'
        ws2[f'D{row}'].fill = calc_fill
        ws2[f'D{row}'].border = border
        ws2[f'D{row}'].number_format = '#,##0.0'

        # Status
        ws2[f'E{row}'] = f'=IF(D{row}<=1,"Excelente",IF(D{row}<=2,"Bom",IF(D{row}<=4,"Aceitavel","Arriscado")))'
        ws2[f'E{row}'].fill = calc_fill
        ws2[f'E{row}'].border = border

    ws2['A16'] = "*Clientes = Break-even / Volume típico (25.000 wf/mês)"
    ws2['A16'].font = Font(italic=True, size=9)

    # Verificação
    ws2['A18'] = "VERIFICAÇÃO (do PDF):"
    ws2['A18'].font = Font(bold=True, color="666666")
    ws2['A19'] = "R$ 0.05: ~94.057 wf, ~3.8 clientes"
    ws2['A19'].font = Font(size=9, color="666666")
    ws2['A20'] = "R$ 0.08: ~58.249 wf, ~2.3 clientes"
    ws2['A20'].font = Font(size=9, color="666666")
    ws2['A21'] = "R$ 0.10: ~46.457 wf, ~1.9 clientes"
    ws2['A21'].font = Font(size=9, color="666666")

    # =========================================================================
    # ABA: SIMULADOR
    # =========================================================================
    ws3 = wb.create_sheet("SIMULADOR")

    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 30

    ws3['A1'] = "SIMULADOR DE MARGEM (Standalone)"
    ws3['A1'].font = Font(bold=True, size=16)

    ws3['A2'] = "Infra 100% dedicada - Worst Case"
    ws3['A2'].font = Font(italic=True)

    # Inputs
    ws3['A4'] = "INPUTS"
    ws3['A4'].font = header_font
    ws3['A4'].fill = header_fill

    ws3['A5'] = "Preço Base (BRL):"
    ws3['B5'] = 2990
    ws3['B5'].fill = input_fill
    ws3['B5'].border = border
    ws3['B5'].number_format = 'R$ #,##0.00'
    ws3['C5'] = "<-- Ajuste aqui"

    ws3['A6'] = "Preço por Workflow (BRL):"
    ws3['B6'] = 0.08
    ws3['B6'].fill = input_fill
    ws3['B6'].border = border
    ws3['B6'].number_format = 'R$ #,##0.00'
    ws3['C6'] = "<-- Ajuste aqui"

    ws3['A7'] = "Volume Workflows/mês:"
    ws3['B7'] = 25000
    ws3['B7'].fill = input_fill
    ws3['B7'].border = border
    ws3['B7'].number_format = '#,##0'
    ws3['C7'] = "<-- Ajuste aqui"

    # Resultados
    ws3['A9'] = "RESULTADOS"
    ws3['A9'].font = header_font
    ws3['A9'].fill = header_fill

    ws3['A10'] = "Receita Base:"
    ws3['B10'] = '=B5'
    ws3['B10'].fill = calc_fill
    ws3['B10'].number_format = 'R$ #,##0.00'

    ws3['A11'] = "Receita Variável:"
    ws3['B11'] = '=B7*B6'
    ws3['B11'].fill = calc_fill
    ws3['B11'].number_format = 'R$ #,##0.00'

    ws3['A12'] = "RECEITA TOTAL:"
    ws3['A12'].font = Font(bold=True)
    ws3['B12'] = '=B10+B11'
    ws3['B12'].font = Font(bold=True)
    ws3['B12'].fill = calc_fill
    ws3['B12'].number_format = 'R$ #,##0.00'

    ws3['A14'] = "Custo Fixo (infra standalone):"
    ws3['B14'] = '=PREMISSAS!C15'  # <<<< R$ 4.590
    ws3['B14'].fill = calc_fill
    ws3['B14'].number_format = 'R$ #,##0.00'

    ws3['A15'] = "Custo Variável:"
    ws3['B15'] = '=B7*PREMISSAS!B42'  # <<<< Volume * Custo var BRL
    ws3['B15'].fill = calc_fill
    ws3['B15'].number_format = 'R$ #,##0.00'

    ws3['A16'] = "CUSTO TOTAL:"
    ws3['A16'].font = Font(bold=True)
    ws3['B16'] = '=B14+B15'
    ws3['B16'].font = Font(bold=True)
    ws3['B16'].fill = calc_fill
    ws3['B16'].number_format = 'R$ #,##0.00'

    ws3['A18'] = "MARGEM BRUTA (R$):"
    ws3['A18'].font = Font(bold=True, size=14)
    ws3['B18'] = '=B12-B16'
    ws3['B18'].font = Font(bold=True, size=14)
    ws3['B18'].fill = calc_fill
    ws3['B18'].number_format = 'R$ #,##0.00'

    ws3['A19'] = "MARGEM BRUTA (%):"
    ws3['A19'].font = Font(bold=True, size=14)
    ws3['B19'] = '=IF(B12>0,B18/B12,0)'
    ws3['B19'].font = Font(bold=True, size=14)
    ws3['B19'].fill = calc_fill
    ws3['B19'].number_format = '0.0%'

    ws3['A21'] = "STATUS:"
    ws3['A21'].font = Font(bold=True, size=14)
    ws3['B21'] = '=IF(B18>0,"SUSTENTAVEL","PREJUIZO")'
    ws3['B21'].font = Font(bold=True, size=14)
    ws3['B21'].fill = calc_fill

    # Verificação
    ws3['A23'] = "VERIFICAÇÃO (do PDF - Growth 25K wf):"
    ws3['A23'].font = Font(bold=True, color="666666")
    ws3['A24'] = "Receita esperada: R$ 4.990"
    ws3['A24'].font = Font(size=9, color="666666")
    ws3['A25'] = "Custo esperado: R$ 4.620"
    ws3['A25'].font = Font(size=9, color="666666")
    ws3['A26'] = "Margem esperada: R$ 370 (7.4%)"
    ws3['A26'].font = Font(size=9, color="666666")

    # =========================================================================
    # ABA: CONFIANCA
    # =========================================================================
    ws4 = wb.create_sheet("CONFIANCA")

    ws4.column_dimensions['A'].width = 40
    ws4.column_dimensions['B'].width = 15
    ws4.column_dimensions['C'].width = 50

    ws4['A1'] = "GRAUS DE CONFIANÇA"
    ws4['A1'].font = Font(bold=True, size=16)

    ws4['A3'] = "Componente"
    ws4['B3'] = "Confiança"
    ws4['C3'] = "Justificativa"
    for col in ['A', 'B', 'C']:
        ws4[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws4[f'{col}3'].fill = header_fill

    items = [
        (4, "Metodologia standalone-first", "ALTO", "Princípio de negócio conservador"),
        (5, "Componentes de infraestrutura", "ALTO", "Baseado em docker-compose.yml real"),
        (6, "Preços de cloud (unitários)", "ALTO", "Pricing público AWS/MongoDB/Temporal"),
        (7, "Sizing de produção", "BAIXO-MEDIO", "Sem Terraform - estimar conservador"),
        (8, "Operações por workflow", "BAIXO-MEDIO", "Baseado em código, não em métricas"),
        (9, "Custo variável por workflow", "BAIXO-MEDIO", "Requer validação com benchmark"),
        (10, "Modelo de pricing (estrutura)", "ALTO", "Padrão de mercado (base + variável)"),
        (11, "Modelo de pricing (valores)", "MEDIO", "Depende de validação de custos"),
    ]

    for row, comp, conf, just in items:
        ws4[f'A{row}'] = comp
        ws4[f'B{row}'] = conf
        ws4[f'C{row}'] = just

    ws4['A13'] = "AÇÕES PARA AUMENTAR CONFIANÇA:"
    ws4['A13'].font = Font(bold=True)

    acoes = [
        (14, "1. Executar benchmark script com Docker"),
        (15, "2. Instalar KubeCost no cluster K8s"),
        (16, "3. Rodar load test com 25K-100K workflows"),
        (17, "4. Criar Terraform do Flowker"),
        (18, "5. Coletar 30+ dias de dados de billing"),
    ]

    for row, acao in acoes:
        ws4[f'A{row}'] = acao

    # =========================================================================
    # VALIDAÇÃO FINAL - Print de verificação
    # =========================================================================
    print("=" * 60)
    print("VALIDAÇÃO DA PLANILHA v3")
    print("=" * 60)

    # Valores esperados
    cambio = 5.00
    custos_cloud = [108, 389, 94, 25, 180, 122]
    total_fixo_usd = sum(custos_cloud)
    total_fixo_brl = total_fixo_usd * cambio

    ops = [(9, 0.000025), (6, 0.000001), (12, 0.0000001)]
    custo_var_usd = sum(qty * cost for qty, cost in ops)
    custo_var_brl = custo_var_usd * cambio

    print(f"Câmbio: {cambio}")
    print(f"Total Fixo USD: ${total_fixo_usd} (esperado: $918)")
    print(f"Total Fixo BRL: R$ {total_fixo_brl:,.2f} (esperado: R$ 4.590)")
    print(f"Custo Var USD: ${custo_var_usd:.6f} (esperado: ~$0.000232)")
    print(f"Custo Var BRL: R$ {custo_var_brl:.6f} (esperado: ~R$ 0.00116)")

    # Break-even para R$ 0.08
    preco = 0.08
    margem = preco - custo_var_brl
    breakeven = total_fixo_brl / margem
    clientes = breakeven / 25000
    print(f"\nBreak-even R$ 0.08:")
    print(f"  Margem: R$ {margem:.4f}")
    print(f"  Break-even: {breakeven:,.0f} wf")
    print(f"  Clientes: {clientes:.1f}")

    # Simulação Growth
    base = 2990
    per_wf = 0.08
    vol = 25000
    receita = base + (vol * per_wf)
    custo = total_fixo_brl + (vol * custo_var_brl)
    margem_sim = receita - custo
    margem_pct = margem_sim / receita
    print(f"\nSimulação Growth (25K wf):")
    print(f"  Receita: R$ {receita:,.2f} (esperado: R$ 4.990)")
    print(f"  Custo: R$ {custo:,.2f} (esperado: R$ 4.620)")
    print(f"  Margem: R$ {margem_sim:,.2f} ({margem_pct:.1%})")
    print("=" * 60)

    # Save
    output_path = "/tmp/pricing-flowker/Flowker_Pricing_Model_v3.xlsx"
    wb.save(output_path)
    print(f"\nPlanilha salva: {output_path}")
    return output_path

if __name__ == "__main__":
    create_pricing_spreadsheet()
